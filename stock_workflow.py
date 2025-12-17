#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台灣股市資料完整處理流程 - GitHub Actions 版本
整合爬蟲、分析、圖表生成的完整自動化流程

作者: Frank
版本: 2.0 (GitHub Actions)
功能:
1. 爬取上市/上櫃每日交易與三大法人資料
2. 清理舊的 History 資料夾
3. 生成分析報告 (Excel) - 分別處理 TSE 和 OTC
4. 清理舊的圖表資料夾
5. 生成技術分析圖表 (HTML + PNG) - 分別處理 TSE 和 OTC
"""

import os
import glob
import shutil
import requests
import pandas as pd
import numpy as np
import time
from datetime import datetime, timedelta
from io import StringIO
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import argparse

# ============================================================================
# 全域設定
# ============================================================================

# 控制是否只分析熱門股票 (買超前150 + 賣超前50)
# True:  只分析買超前150 + 賣超前50
# False: 分析所有 CSV 內的股票
TOP_STOCKS_ONLY = True
<<<<<<< HEAD

=======
>>>>>>> 3c0490187701b8052424463ee13effe7be83bfec
# ============================================================================
# 共用工具函數
# ============================================================================

def clean_excel_keep_second_sheet(input_file):
    """
    只保留 Excel 的第二個分頁（最近交易日），並以該分頁的日期重新命名檔案
    
    Args:
        input_file: 輸入的 Excel 檔案路徑
        
    Returns:
        新檔案路徑或 None (如果失敗)
    """
    
    # 檢查檔案是否存在
    if not os.path.exists(input_file):
        print(f"  ❌ 檔案不存在: {input_file}")
        return None
    
    try:
        # 載入 Excel 檔案
        wb = load_workbook(input_file)
        sheet_names = wb.sheetnames
        
        print(f"  📋 原始分頁數: {len(sheet_names)}")
        
        # 檢查是否至少有 2 個分頁
        if len(sheet_names) < 2:
            print(f"  ⚠️  只有 {len(sheet_names)} 個分頁，跳過清理")
            wb.close()
            return None
        
        # 取得第二個分頁的名稱（這是要保留的）
        second_sheet_name = sheet_names[1]
        print(f"  ✓ 保留分頁: {second_sheet_name}")
        
        # 從分頁名稱提取日期 (假設格式為 YYYYMMDD)
        match = re.search(r'(\d{8})', second_sheet_name)
        if not match:
            print(f"  ⚠️  無法從分頁名稱提取日期: {second_sheet_name}")
            wb.close()
            return None
        
        new_date_str = match.group(1)
        print(f"  📅 提取日期: {new_date_str}")
        
        # 刪除其他所有分頁（除了第二個）
        sheets_to_delete = [name for i, name in enumerate(sheet_names) if i != 1]
        for sheet_name in sheets_to_delete:
            wb.remove(wb[sheet_name])
            print(f"  🗑️  已刪除分頁: {sheet_name}")
        
        # 生成新檔案名稱
        dir_name = os.path.dirname(input_file)
        base_name = os.path.basename(input_file)
        
        # 提取檔案前綴 (tse_analysis_result 或 otc_analysis_result)
        if 'tse_analysis_result' in base_name.lower():
            prefix = 'tse_analysis_result'
        elif 'otc_analysis_result' in base_name.lower():
            prefix = 'otc_analysis_result'
        else:
            # 使用原始檔案名去掉日期部分
            prefix = re.sub(r'_\d{8}', '', base_name.replace('.xlsx', ''))
        
        new_file_name = f"{prefix}_{new_date_str}.xlsx"
        new_file_path = os.path.join(dir_name, new_file_name)
        
        # 儲存新檔案
        wb.save(new_file_path)
        wb.close()
        
        file_size = os.path.getsize(new_file_path) / 1024  # KB
        print(f"  ✅ 已生成: {new_file_name} ({file_size:.1f} KB)")
        
        # 如果新舊檔案名不同，刪除舊檔案
        if new_file_path != input_file:
            try:
                os.remove(input_file)
                print(f"  🗑️  已刪除舊檔: {base_name}")
            except Exception as e:
                print(f"  ⚠️  無法刪除舊檔: {e}")
        
        return new_file_path
        
    except Exception as e:
        print(f"  ❌ 處理失敗: {e}")
        import traceback
        traceback.print_exc()
        return None


def setup_base_directory():
    """
    設定基礎工作目錄
    在 GitHub Actions 中使用當前工作目錄或指定的資料目錄
    """
    # 優先使用環境變數指定的目錄
    base_dir = os.environ.get('STOCK_DATA_DIR', os.getcwd())
    
    # 確保目錄存在
    os.makedirs(base_dir, exist_ok=True)
    
    print(f"✓ 工作目錄: {base_dir}\n")
    return base_dir

def delete_folders(base_dir, folder_names):
    """刪除並重建指定的資料夾"""
    print(f"\n{'='*80}")
    print("清理資料夾...")
    print(f"{'='*80}")
    
    for folder_name in folder_names:
        folder_path = os.path.join(base_dir, folder_name)
        
        # 統計現有檔案數量
        file_count = 0
        if os.path.exists(folder_path):
            try:
                files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
                file_count = len(files)
                print(f"📂 {folder_name}: 發現 {file_count} 個 CSV 檔案")
            except Exception as e:
                print(f"⚠️  無法讀取 {folder_name}: {e}")
        
        # 刪除資料夾
        if os.path.exists(folder_path):
            try:
                shutil.rmtree(folder_path)
                print(f"✓ 已刪除: {folder_name} ({file_count} 個檔案)")
                
                # 等待檔案系統完成操作
                import time
                time.sleep(0.5)
                
            except Exception as e:
                print(f"✗ 刪除失敗 {folder_name}: {e}")
                continue
        else:
            print(f"⊘ 資料夾不存在: {folder_name}")
        
        # 重新建立空資料夾
        try:
            os.makedirs(folder_path, exist_ok=True)
            print(f"✓ 已重建空資料夾: {folder_name}")
            
            # 驗證資料夾是空的
            remaining = os.listdir(folder_path)
            if remaining:
                print(f"⚠️  警告: {folder_name} 內還有 {len(remaining)} 個項目！")
            
        except Exception as e:
            print(f"✗ 重建資料夾失敗 {folder_name}: {e}")
    
    print(f"{'='*80}\n")

def create_required_directories(base_dir):
    """建立所需的資料夾結構"""
    required_dirs = [
        'StockList',       # 股票清單和字體
        'StockTSEDaily',
        'StockTSEShares',
        'StockOTCDaily',
        'StockOTCShares',
        'StockInfo',       # 分析報告
        'StockTSEHistory',
        'StockOTCHistory',
        'StockTSEHTML',
        'StockOTCHTML',
        'local_StockTSEHistory',  # 新增 local 資料夾
        'local_StockOTCHistory',
        'local_StockTSEHTML',
        'local_StockOTCHTML'
    ]
    
    print(f"\n{'='*80}")
    print("建立資料夾結構...")
    print(f"{'='*80}")
    
    for dir_name in required_dirs:
        dir_path = os.path.join(base_dir, dir_name)
        os.makedirs(dir_path, exist_ok=True)
        print(f"✓ {dir_name}")
    
    print(f"{'='*80}\n")

# ============================================================================
# 第一步：爬蟲程式的所有函數
# ============================================================================

# 【第一步-filter_csv_content】
# 從第一步程式複製 filter_csv_content 函數
def filter_csv_content(csv_bytes):
    """過濾 CSV 內容，只保留股票資料"""
    try:
        content = csv_bytes.decode('cp950')
        lines = content.split('\r\n')

        filtered_lines = []
        header_found = False
        stock_count = 0

        for line in lines:
            if '證券代號' in line and not header_found:
                filtered_lines.append(line)
                header_found = True
                continue

            if header_found:
                match = re.match(r'^=?"?(\d{4})"?', line)
                if match:
                    filtered_lines.append(line)
                    stock_count += 1

        filtered_content = '\r\n'.join(filtered_lines)
        filtered_bytes = filtered_content.encode('cp950')
        print(f"   ✂️  過濾完成：保留 {stock_count} 檔股票")
        return filtered_bytes

    except Exception as e:
        print(f"   ⚠️  過濾失敗: {e}，將儲存原始資料")
        return csv_bytes

# 【第一步-download_twse_daily】
# 從第一步程式複製 download_twse_daily 函數
def download_twse_daily(date_str):
    """下載上市每日交易資料"""
    if '-' in date_str:
        date_str = date_str.replace('-', '')

    url = f"https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date={date_str}&type=ALL&response=csv"

    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200 and len(response.content) > 100:
            return response.content
        return None
    except Exception as e:
        print(f"   ❌ 下載錯誤: {e}")
        return None
# 【第一步-crawl_twse_daily】
# 從第一步程式複製 crawl_twse_daily 函數
def crawl_twse_daily(start_date, end_date, save_dir):
    """抓取上市每日交易資料"""
    print("="*60)
    print("📊 [1/4] 上市每日交易資料 (TWSE Daily)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    # 從今天往回檢查
    while curr >= start_date:
        if curr.weekday() < 5:  # 只檢查平日
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                print(f"  {date_formatted}... [已存在，停止檢查] ✓")
                break
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        csv_bytes = download_twse_daily(date_str)

        if csv_bytes:
            filtered_bytes = filter_csv_content(csv_bytes)
            with open(file_path, 'wb') as f:
                f.write(filtered_bytes)
            print(" ✓")
            success_count += 1
        else:
            print(" ✗")

        time.sleep(1)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count
# 【第一步-download_twse_institutional】
# 從第一步程式複製 download_twse_institutional 函數
def download_twse_institutional(date_str):
    """下載上市三大法人資料"""
    url = 'https://www.twse.com.tw/rwd/zh/fund/T86'
    params = {'date': date_str, 'selectType': 'ALL', 'response': 'json'}
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    try:
        response = requests.get(url, params=params, headers=headers, timeout=30)
        response.raise_for_status()
        data = response.json()

        if data.get('stat') == 'OK' and 'data' in data:
            return pd.DataFrame(data['data'], columns=data['fields'])
        return None
    except Exception as e:
        print(f"   ❌ 錯誤: {e}")
        return None
# 【第一步-crawl_twse_institutional】
# 從第一步程式複製 crawl_twse_institutional 函數
def crawl_twse_institutional(start_date, end_date, save_dir):
    """抓取上市三大法人買賣超資料"""
    print("="*60)
    print("📊 [2/4] 上市三大法人買賣超 (TWSE Institutional)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    while curr >= start_date:
        if curr.weekday() < 5:
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                print(f"  {date_formatted}... [已存在，停止檢查] ✓")
                break
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        df = download_twse_institutional(date_str)

        if df is not None and not df.empty:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(" ✓")
            success_count += 1
        else:
            print(" ✗")

        time.sleep(3)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count
# 【第一步-process_otc_daily_columns】
# 從第一步程式複製 process_otc_daily_columns 函數
def process_otc_daily_columns(df):
    """處理上櫃每日交易資料欄位"""
    rename_mapping = {
        '代號': '證券代號',
        '名稱': '證券名稱',
        '收盤': '收盤價',
        '開盤': '開盤價',
        '最高': '最高價',
        '最低': '最低價',
        '成交股數': '成交股數',
        '成交筆數': '成交筆數',
        '成交金額(元)': '成交金額',
        '漲跌': '漲跌價差',
        '最後買價': '最後揭示買價',
        '最後買量(千股)': '最後揭示買量',
        '最後賣價': '最後揭示賣價',
        '最後賣量(千股)': '最後揭示賣量'
    }

    df = df.rename(columns=rename_mapping)

    # 刪除不需要的欄位
    columns_to_drop = ['均價', '發行股數', '次日參考價', '次日漲停價', '次日跌停價']
    existing_cols_to_drop = [col for col in columns_to_drop if col in df.columns]
    if existing_cols_to_drop:
        df = df.drop(columns=existing_cols_to_drop)

    # 新增漲跌(+/-)欄位
    if '漲跌價差' in df.columns:
        df['漲跌價差'] = pd.to_numeric(df['漲跌價差'], errors='coerce')
        df['漲跌(+/-)'] = df['漲跌價差'].apply(lambda x: '+' if x > 0 else '-' if pd.notna(x) else '')
        df['漲跌價差'] = df['漲跌價差'].abs()
    else:
        df['漲跌(+/-)'] = ''

    # 新增本益比欄位
    df['本益比'] = ''

    # 調整欄位順序
    desired_order = [
        '證券代號', '證券名稱', '成交股數', '成交筆數', '成交金額',
        '開盤價', '最高價', '最低價', '收盤價', '漲跌(+/-)', '漲跌價差',
        '最後揭示買價', '最後揭示買量', '最後揭示賣價', '最後揭示賣量', '本益比'
    ]

    existing_desired_cols = [col for col in desired_order if col in df.columns]
    other_cols = [col for col in df.columns if col not in desired_order]
    final_order = existing_desired_cols + other_cols
    df = df[final_order]

    return df
# 【第一步-download_otc_daily】
# 從第一步程式複製 download_otc_daily 函數
def download_otc_daily(date_str):
    """下載上櫃每日交易資料"""
    date_formatted = f"{date_str[:4]}%2F{date_str[4:6]}%2F{date_str[6:]}"
    url = f'https://www.tpex.org.tw/www/zh-tw/afterTrading/dailyQuotes?date={date_formatted}&id=&response=csv'

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Referer': 'https://www.tpex.org.tw/zh-tw/aftertrading/quotes/daily.html'
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()

        if not response.content or len(response.content) < 100:
            return None

        encodings = ['big5', 'cp950', 'utf-8', 'utf-8-sig']

        for encoding in encodings:
            try:
                text = response.content.decode(encoding)

                if '查無資料' in text or '目前無資料' in text:
                    return None

                csv_data = StringIO(text)
                df = pd.read_csv(csv_data, skiprows=2)

                if df.empty:
                    continue

                df = df.dropna(how='all')

                if len(df.columns) > 0:
                    first_col = df.columns[0]
                    df = df[df[first_col].notna()]
                    df = df[~df[first_col].astype(str).str.contains('上櫃|總成交|註:', na=False)]

                if len(df) == 0:
                    continue

                first_col = df.columns[0] if len(df.columns) > 0 else ''
                if any('\u4e00' <= c <= '\u9fff' for c in first_col):
                    df = process_otc_daily_columns(df)
                    return df

            except:
                continue

        return None

    except Exception as e:
        return None
# 【第一步-crawl_otc_daily】
# 從第一步程式複製 crawl_otc_daily 函數
def crawl_otc_daily(start_date, end_date, save_dir):
    """抓取上櫃每日交易資料"""
    print("="*60)
    print("📊 [3/4] 上櫃每日交易資料 (OTC Daily)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    while curr >= start_date:
        if curr.weekday() < 5:
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                try:
                    df_check = pd.read_csv(file_path)
                    if len(df_check) > 1:
                        print(f"  {date_formatted}... [已存在，停止檢查] ✓")
                        break
                    else:
                        missing_dates.append(curr)
                except:
                    missing_dates.append(curr)
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        df = download_otc_daily(date_str)

        if df is not None and not df.empty:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(f" ✓ ({len(df)} 筆)")
            success_count += 1
        else:
            print(" ✗")

        if idx % 5 == 0:
            time.sleep(4)
        else:
            time.sleep(2)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count

# 【第一步-process_otc_institutional_columns】
# 從第一步程式複製 process_otc_institutional_columns 函數
def process_otc_institutional_columns(df):
    """處理上櫃三大法人資料欄位"""
    column_rename_map = {
        '代號': '證券代號',
        '名稱': '證券名稱',
        '外資及陸資(不含外資自營商)-買進股數': '外陸資買進股數(不含外資自營商)',
        '外資及陸資(不含外資自營商)-賣出股數': '外陸資賣出股數(不含外資自營商)',
        '外資及陸資(不含外資自營商)-買賣超股數': '外陸資買賣超股數(不含外資自營商)',
        '外資自營商-買進股數': '外資自營商買進股數',
        '外資自營商-賣出股數': '外資自營商賣出股數',
        '外資自營商-買賣超股數': '外資自營商買賣超股數',
        '投信-買進股數': '投信買進股數',
        '投信-賣出股數': '投信賣出股數',
        '投信-買賣超股數': '投信買賣超股數',
        '自營商(自行買賣)-買進股數': '自營商買進股數(自行買賣)',
        '自營商(自行買賣)-賣出股數': '自營商賣出股數(自行買賣)',
        '自營商(自行買賣)-買賣超股數': '自營商買賣超股數(自行買賣)',
        '自營商(避險)-買進股數': '自營商買進股數(避險)',
        '自營商(避險)-賣出股數': '自營商賣出股數(避險)',
        '自營商(避險)-買賣超股數': '自營商買賣超股數(避險)',
        '自營商-買賣超股數': '自營商買賣超股數',
        '三大法人買賣超股數合計': '三大法人買賣超股數'
    }

    df = df.rename(columns=column_rename_map)

    # 刪除指定欄位
    columns_to_drop_indices = [8, 9, 10, 20, 21]
    all_columns = list(df.columns)
    columns_to_keep = [col for idx, col in enumerate(all_columns) if idx not in columns_to_drop_indices]
    df = df[columns_to_keep]

    # 調整欄位順序
    current_columns = list(df.columns)
    if '自營商買賣超股數' in current_columns and '投信買賣超股數' in current_columns:
        current_columns.remove('自營商買賣超股數')
        invest_trust_idx = current_columns.index('投信買賣超股數')
        current_columns.insert(invest_trust_idx + 1, '自營商買賣超股數')
        df = df[current_columns]

    return df

# 【第一步-download_otc_institutional】
# 從第一步程式複製 download_otc_institutional 函數
def download_otc_institutional(date_str):
    """下載上櫃三大法人資料"""
    date_formatted = f"{date_str[:4]}%2F{date_str[4:6]}%2F{date_str[6:]}"
    url = f'https://www.tpex.org.tw/www/zh-tw/insti/dailyTrade?type=Daily&sect=AL&date={date_formatted}&id=&response=csv'

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Referer': 'https://www.tpex.org.tw/zh-tw/mainboard/trading/major-institutional/detail/day.html'
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()

        if not response.content or len(response.content) < 100:
            return None

        encodings = ['big5', 'cp950', 'utf-8', 'utf-8-sig']

        for encoding in encodings:
            try:
                text = response.content.decode(encoding)

                if '查無資料' in text or '目前無資料' in text:
                    return None

                csv_data = StringIO(text)
                df = pd.read_csv(csv_data, skiprows=1)

                if df.empty or len(df) == 0:
                    continue

                df = df.dropna(how='all')

                if len(df) == 0:
                    continue

                first_col = df.columns[0] if len(df.columns) > 0 else ''
                if any('\u4e00' <= c <= '\u9fff' for c in first_col):
                    df = process_otc_institutional_columns(df)
                    return df

            except:
                continue

        return None

    except Exception as e:
        return None
# 【第一步-crawl_otc_institutional】
# 從第一步程式複製 crawl_otc_institutional 函數
def crawl_otc_institutional(start_date, end_date, save_dir):
    """抓取上櫃三大法人買賣超資料"""
    print("="*60)
    print("📊 [4/4] 上櫃三大法人買賣超 (OTC Institutional)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    while curr >= start_date:
        if curr.weekday() < 5:
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                try:
                    df_check = pd.read_csv(file_path)
                    if len(df_check) > 1:
                        print(f"  {date_formatted}... [已存在，停止檢查] ✓")
                        break
                    else:
                        missing_dates.append(curr)
                except:
                    missing_dates.append(curr)
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        df = download_otc_institutional(date_str)

        if df is not None and not df.empty:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(f" ✓ ({len(df)} 筆)")
            success_count += 1
        else:
            print(" ✗")

        if idx % 5 == 0:
            time.sleep(4)
        else:
            time.sleep(2)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count

def run_step1_crawler(base_dir, start_date=None, end_date=None):
    """執行第一步：爬蟲程式"""
    print("\n" + "🔥"*40)
    print("第一步：執行爬蟲程式")
    print("🔥"*40 + "\n")
    
    if start_date is None:
        start_date = datetime(2025, 1, 1)
    if end_date is None:
        end_date = datetime.now()

    print(f"日期範圍: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
    print(f"儲存位置: {base_dir}/")
    print()

    start_time = time.time()

    dirs = {
        'StockTSEDaily': os.path.join(base_dir, 'StockTSEDaily'),
        'StockTSEShares': os.path.join(base_dir, 'StockTSEShares'),
        'StockOTCDaily': os.path.join(base_dir, 'StockOTCDaily'),
        'StockOTCShares': os.path.join(base_dir, 'StockOTCShares')
    }

    results = {}
    results['twse_daily'] = crawl_twse_daily(start_date, end_date, dirs['StockTSEDaily'])
    results['twse_inst'] = crawl_twse_institutional(start_date, end_date, dirs['StockTSEShares'])
    results['otc_daily'] = crawl_otc_daily(start_date, end_date, dirs['StockOTCDaily'])
    results['otc_inst'] = crawl_otc_institutional(start_date, end_date, dirs['StockOTCShares'])

    elapsed_time = time.time() - start_time

    print("="*60)
    print("📊 第一步執行結果摘要")
    print("="*60)
    print(f"✓ 上市每日交易：  {results['twse_daily']} 個檔案")
    print(f"✓ 上市三大法人：  {results['twse_inst']} 個檔案")
    print(f"✓ 上櫃每日交易：  {results['otc_daily']} 個檔案")
    print(f"✓ 上櫃三大法人：  {results['otc_inst']} 個檔案")
    print("-"*60)
    print(f"總計下載：{sum(results.values())} 個檔案")
    print(f"執行時間：{elapsed_time:.1f} 秒")
    print("="*60)

# ============================================================================
# 第二步：分析程式的所有函數
# ============================================================================

# 【第二步-normalize_stock_code】
# 從第二步程式複製 normalize_stock_code 函數
def normalize_stock_code(code):
    """
    標準化股票代碼，確保像'56'會被轉換成'0056'
    規則：如果是純數字且長度小於4，則補0到4位數
    """
    if pd.isna(code) or code == '':
        return ''

    code_str = str(code).strip()
    code_str = code_str.replace('="', '').replace('"', '').replace("'", '')

    if code_str.isdigit() and len(code_str) < 4:
        return code_str.zfill(4)

    return code_str

# 【第二步-shares_to_lots】
# 從第二步程式複製 shares_to_lots 函數
def shares_to_lots(value):
    """
    將股數轉換為張數(除以1000後取整數)
    小於1000股視為0張
    """
    try:
        if pd.isna(value) or value == '':
            return 0
        if isinstance(value, str):
            value = value.replace(',', '')
        num_value = float(value)
        return int(num_value / 1000)
    except:
        return 0
    
# 【第二步-format_date_short】
# 從第二步程式複製 format_date_short 函數
def format_date_short(date_str):
    """將 YYYY-MM-DD 格式轉換為 DD (只顯示日)"""
    try:
        parts = date_str.split('-')
        if len(parts) == 3:
            return f"{parts[2]}"
        return date_str
    except:
        return date_str
    
# 【第二步-setup_config】
# 從第二步程式複製 setup_config 函數 (需要修改路徑)
def setup_config(market_type='TSE'):
    """
    設定所有路徑變數 (GitHub Actions 版本)

    Args:
        market_type: 'TSE' (上市) 或 'OTC' (上櫃)

    Returns:
        dict: 包含所有路徑配置的字典
    """
    # GitHub Actions 使用當前目錄
    base_path = os.getcwd()

    if market_type == 'TSE':
        config = {
            'market_type': market_type,
            'folder_path': os.path.join(base_path, 'StockTSEShares'),
            'stock_daily_folder': os.path.join(base_path, 'StockTSEDaily'),
            'output_folder': os.path.join(base_path, 'StockInfo'),
            'history_folder': os.path.join(base_path, 'StockTSEHistory'),
            'market_list_filename': 'tse_company_list.csv',
            'output_filename': 'tse_analysis_result.xlsx',
            'sigma_threshold': 2.5,
            'aggregate_threshold': None,
            'show_top_n': 100,
            'top_buy_count': 100,   # 買超前n名
            'top_sell_count': 50   # 賣超前n名
        }
    else:  # OTC
        config = {
            'market_type': market_type,
            'folder_path': os.path.join(base_path, 'StockOTCShares'),
            'stock_daily_folder': os.path.join(base_path, 'StockOTCDaily'),
            'output_folder': os.path.join(base_path, 'StockInfo'),
            'history_folder': os.path.join(base_path, 'StockOTCHistory'),
            'market_list_filename': 'otc_company_list.csv',
            'output_filename': 'otc_analysis_result.xlsx',
            'sigma_threshold': 2.5,
            'aggregate_threshold': None,
            'show_top_n': 100,
            'top_buy_count': 100,   # 買超前n名
            'top_sell_count': 50   # 賣超前n名
        }

    # 建立完整路徑
    config['market_list_path'] = os.path.join(config['output_folder'], config['market_list_filename'])
    config['output_path'] = os.path.join(config['output_folder'], config['output_filename'])

    # 建立輸出資料夾
    os.makedirs(config['output_folder'], exist_ok=True)
    os.makedirs(config['history_folder'], exist_ok=True)

    print(f"{'='*80}")
    print(f"市場類型: {market_type} ({'上市' if market_type == 'TSE' else '上櫃'})")
    print(f"三大法人資料夾: {config['folder_path']}")
    print(f"個股日線資料夾: {config['stock_daily_folder']}")
    print(f"輸出資料夾: {config['output_folder']}")
    print(f"歷史數據資料夾: {config['history_folder']}")
    print(f"股票清單檔案: {config['market_list_path']}")
    print(f"輸出Excel檔案: {config['output_path']}")
    print(f"買超分析數量: 前 {config['top_buy_count']} 名")
    print(f"賣超分析數量: 前 {config['top_sell_count']} 名")
    if config['show_top_n'] is not None:
        print(f"彙整分析模式: 顯示前 {config['show_top_n']} 名")
    else:
        print(f"彙整分析閾值: {config['aggregate_threshold']} 張")
    print(f"{'='*80}\n")

    return config
# 【第二步-load_stock_list】
# 從第二步程式複製 load_stock_list 函數
def load_stock_list(market_list_path):
    """
    讀取允許的股票代碼清單和領域資訊

    Returns:
        tuple: (allowed_stock_codes, stock_sector_map, etf_stock_codes)
    """
    allowed_stock_codes = set()
    stock_sector_map = {}
    etf_stock_codes = set()

    try:
        market_df = pd.read_csv(market_list_path, encoding='utf-8')
        first_column = market_df.iloc[:, 0].apply(normalize_stock_code)
        allowed_stock_codes = set(first_column.tolist())

        if len(market_df.columns) >= 3:
            for idx, row in market_df.iterrows():
                stock_code = normalize_stock_code(row.iloc[0])
                sector = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                stock_sector_map[stock_code] = sector

                if sector.upper() == 'ETF':
                    etf_stock_codes.add(stock_code)

            print(f"{'='*80}")
            print(f"已載入允許的股票代碼清單: {len(allowed_stock_codes)} 檔")
            print(f"已建立領域映射: {len(stock_sector_map)} 檔")
            print(f"識別ETF股票: {len(etf_stock_codes)} 檔")
            print(f"清單檔案: {market_list_path}")
            print(f"前10個代碼: {list(allowed_stock_codes)[:10]}")
            if etf_stock_codes:
                print(f"ETF代碼: {sorted(etf_stock_codes)}")
            print(f"{'='*80}\n")
        else:
            print(f"{'='*80}")
            print(f"已載入允許的股票代碼清單: {len(allowed_stock_codes)} 檔")
            print(f"警告: 欄位不足，無法讀取領域資訊（需要至少3欄）")
            print(f"{'='*80}\n")

    except FileNotFoundError:
        print(f"警告: 找不到 {market_list_path}")
        print("將處理所有股票代碼")
        return None, {}, set()
    except Exception as e:
        print(f"讀取股票清單時發生錯誤: {e}")
        print("將處理所有股票代碼")
        return None, {}, set()

    return allowed_stock_codes, stock_sector_map, etf_stock_codes


# 【第二步-is_allowed_stock】
# 從第二步程式複製 is_allowed_stock 函數
def is_allowed_stock(stock_code, allowed_stock_codes):
    """檢查股票代碼是否在允許清單中"""
    if allowed_stock_codes is None:
        return True
    normalized_code = normalize_stock_code(stock_code)
    return normalized_code in allowed_stock_codes

# 【第二步-get_stock_sector】
# 從第二步程式複製 get_stock_sector 函數
def get_stock_sector(stock_code, stock_sector_map):
    """獲取股票代碼對應的領域"""
    normalized_code = normalize_stock_code(stock_code)
    return stock_sector_map.get(normalized_code, '')

# 【第二步-load_stock_daily_prices】
# 從第二步程式複製 load_stock_daily_prices 函數
def load_stock_daily_prices(stock_daily_folder, allowed_stock_codes, num_days=5):
    """
    讀取StockTSEDaily的收盤價和漲跌價差

    Returns:
        dict: {日期: {證券代號: {'收盤價': x, '漲跌價差': y}}}
    """
    stock_daily_prices = {}

    print(f"\n{'='*80}")
    print("開始讀取 StockTSEDaily 的收盤價和漲跌價差資料...")
    print(f"{'='*80}")

    if not os.path.exists(stock_daily_folder):
        print(f"警告: StockTSEDaily 資料夾不存在: {stock_daily_folder}")
        print("將無法顯示收盤價和漲跌價差")
        print(f"{'='*80}\n")
        return stock_daily_prices

    all_daily_files = glob.glob(os.path.join(stock_daily_folder, '*.csv'))
    daily_files_sorted = sorted(all_daily_files, key=lambda x: os.path.basename(x).replace('.csv', ''), reverse=True)
    latest_files = daily_files_sorted[:num_days]

    print(f"找到 {len(all_daily_files)} 個 StockTSEDaily 檔案")
    print(f"將讀取最近 {num_days} 個檔案的價格資料")

    for daily_file in latest_files:
        try:
            # 先嘗試 cp950 編碼,失敗則用 utf-8
            try:
                df_daily = pd.read_csv(daily_file, encoding='cp950', low_memory=False)
            except:
                df_daily = pd.read_csv(daily_file, encoding='utf-8', low_memory=False)

            file_date = os.path.basename(daily_file).replace('.csv', '')

            if '證券代號' in df_daily.columns:
                df_daily['證券代號'] = df_daily['證券代號'].apply(normalize_stock_code)

            if allowed_stock_codes is not None:
                df_daily = df_daily[df_daily['證券代號'].isin(allowed_stock_codes)]

            stock_daily_prices[file_date] = {}

            for _, row in df_daily.iterrows():
                stock_code = normalize_stock_code(row['證券代號'])
                close_price = row.get('收盤價', '')

                price_sign = ''
                if len(df_daily.columns) > 9:
                    j_col_name = df_daily.columns[9]
                    price_sign = str(row.get(j_col_name, '')).strip()

                price_value = ''
                if len(df_daily.columns) > 10:
                    k_col_name = df_daily.columns[10]
                    price_value = str(row.get(k_col_name, '')).strip()

                if price_sign and price_value and price_value not in ['', '--', 'X']:
                    clean_value = price_value.replace(',', '')
                    price_diff = f"{price_sign}{clean_value}"
                else:
                    price_diff = ''

                stock_daily_prices[file_date][stock_code] = {
                    '收盤價': close_price,
                    '漲跌價差': price_diff
                }

            print(f"  已讀取: {os.path.basename(daily_file)} - {len(stock_daily_prices[file_date])} 檔股票")

        except Exception as e:
            print(f"讀取StockTSEDaily檔案 {daily_file} 時發生錯誤: {e}")

    print(f"完成讀取價格資料,共 {len(stock_daily_prices)} 天")
    print(f"{'='*80}\n")

    return stock_daily_prices

# 【第二步-get_latest_files】
# 從第二步程式複製 get_latest_files 函數
def get_latest_files(folder_path, num_files=61):
    """取得最新的N個檔案"""
    csv_files = glob.glob(os.path.join(folder_path, '*.csv'))
    csv_files_sorted = sorted(csv_files, key=lambda x: os.path.basename(x).replace('.csv', ''), reverse=True)
    return csv_files_sorted[:num_files]

# 【第二步-process_shares_files】
# 從第二步程式複製 process_shares_files 函數
def process_shares_files(latest_files, allowed_stock_codes, stock_daily_prices,
                         stock_sector_map, etf_stock_codes, top_buy_count=50, top_sell_count=20):
    """
    處理三大法人買賣超檔案

    Args:
        top_buy_count: 買超顯示前N名 (預設50)
        top_sell_count: 賣超顯示前N名 (預設20)

    Returns:
        tuple: (all_data, daily_buy_sell_data, etf_daily_data, buy_top20_tracker,
                sell_top20_tracker, daily_buy_stocks, daily_sell_stocks,
                daily_all_stocks, all_historical_data, statistics)
    """
    all_data = []
    daily_buy_sell_data = []
    etf_daily_data = []
    buy_top20_tracker = []
    sell_top20_tracker = []
    daily_buy_stocks = {}
    daily_sell_stocks = {}
    daily_all_stocks = {}
    all_historical_data = {}

    filtered_out_count = 0
    processed_count = 0

    print(f"找到 {len(latest_files)} 個 CSV 檔案")
    print(f"將處理最新的 {len(latest_files)} 個檔案用於標準差計算")
    print(f"最近5個檔案:")
    for i, file in enumerate(latest_files[:5], 1):
        print(f"{i}. {os.path.basename(file)}")

    for file_path in latest_files:
        try:
            df = pd.read_csv(file_path, encoding='utf-8')

            if '證券代號' in df.columns:
                df['證券代號'] = df['證券代號'].apply(normalize_stock_code)

            if allowed_stock_codes is not None:
                original_count = len(df)
                df = df[df['證券代號'].isin(allowed_stock_codes)]
                filtered_count = original_count - len(df)
                filtered_out_count += filtered_count
                processed_count += len(df)

            file_date = os.path.basename(file_path).replace('.csv', '')

            if '三大法人買賣超股數' in df.columns:
                df['三大法人買賣超股數'] = pd.to_numeric(
                    df['三大法人買賣超股數'].astype(str).str.replace(',', ''),
                    errors='coerce'
                )
                df['買賣超張數'] = (df['三大法人買賣超股數'] / 1000).fillna(0).astype(int)

                # 記錄每天所有股票的買賣超狀態
                daily_all_stocks[file_date] = {}
                for _, row in df.iterrows():
                    if pd.notna(row['證券代號']) and pd.notna(row['買賣超張數']):
                        stock_code = normalize_stock_code(row['證券代號'])
                        if is_allowed_stock(stock_code, allowed_stock_codes):
                            buy_sell_value = int(row['買賣超張數'])
                            daily_all_stocks[file_date][stock_code] = buy_sell_value

                            if stock_code not in all_historical_data:
                                all_historical_data[stock_code] = []
                            all_historical_data[stock_code].append((file_date, buy_sell_value))

                # 只處理前5天的詳細資料
                if file_path in latest_files[:5]:
                    print(f"\n{'='*80}")
                    print(f"檔案:{os.path.basename(file_path)}")
                    print(f"{'='*80}")

                    # 買超處理 - 使用參數控制數量
                    buy_top = df[df['買賣超張數'] > 0].nlargest(top_buy_count, '買賣超張數')
                    print(f"\n【買超 TOP {top_buy_count}】")
                    print("-" * 80)

                    if len(buy_top) > 0:
                        display_df = buy_top[['證券代號', '證券名稱', '買賣超張數']].copy()
                        print(display_df.to_string(index=False))

                        buy_top20 = df[df['買賣超張數'] > 0].nlargest(20, '買賣超張數')
                        daily_buy_stocks[file_date] = set(buy_top20['證券代號'].tolist())

                        buy_output = buy_top[['證券代號', '證券名稱', '買賣超張數']].copy()
                        buy_output['日期'] = file_date
                        buy_output['類別'] = '買超'
                        buy_output['排名'] = range(1, len(buy_output) + 1)

                        if file_date in stock_daily_prices:
                            buy_output['收盤價'] = buy_output['證券代號'].apply(
                                lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('收盤價', '')
                            )
                            buy_output['漲跌價差'] = buy_output['證券代號'].apply(
                                lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('漲跌價差', '')
                            )
                        else:
                            buy_output['收盤價'] = ''
                            buy_output['漲跌價差'] = ''

                        daily_buy_sell_data.append(buy_output)

                        for _, row in buy_top20.iterrows():
                            buy_top20_tracker.append({
                                '證券代號': normalize_stock_code(row['證券代號']),
                                '證券名稱': row['證券名稱'],
                                '日期': file_date,
                                '買賣超張數': int(row['買賣超張數'])
                            })
                    else:
                        print("無買超資料")
                        daily_buy_stocks[file_date] = set()

                    # 賣超處理 - 使用參數控制數量
                    sell_top = df[df['買賣超張數'] < 0].nsmallest(top_sell_count, '買賣超張數')
                    print(f"\n【賣超 TOP {top_sell_count}】")
                    print("-" * 80)

                    if len(sell_top) > 0:
                        display_df = sell_top[['證券代號', '證券名稱', '買賣超張數']].copy()
                        print(display_df.to_string(index=False))

                        sell_top20 = df[df['買賣超張數'] < 0].nsmallest(20, '買賣超張數')
                        daily_sell_stocks[file_date] = set(sell_top20['證券代號'].tolist())

                        sell_output = sell_top[['證券代號', '證券名稱', '買賣超張數']].copy()
                        sell_output['日期'] = file_date
                        sell_output['類別'] = '賣超'
                        sell_output['排名'] = range(1, len(sell_output) + 1)

                        if file_date in stock_daily_prices:
                            sell_output['收盤價'] = sell_output['證券代號'].apply(
                                lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('收盤價', '')
                            )
                            sell_output['漲跌價差'] = sell_output['證券代號'].apply(
                                lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('漲跌價差', '')
                            )
                        else:
                            sell_output['收盤價'] = ''
                            sell_output['漲跌價差'] = ''

                        daily_buy_sell_data.append(sell_output)

                        for _, row in sell_top20.iterrows():
                            sell_top20_tracker.append({
                                '證券代號': normalize_stock_code(row['證券代號']),
                                '證券名稱': row['證券名稱'],
                                '日期': file_date,
                                '買賣超張數': int(row['買賣超張數'])
                            })
                    else:
                        print("無賣超資料")
                        daily_sell_stocks[file_date] = set()

                    # ETF處理
                    if len(etf_stock_codes) > 0:
                        etf_df = df[df['證券代號'].isin(etf_stock_codes)].copy()

                        if len(etf_df) > 0:
                            # ETF買超
                            etf_buy_top10 = etf_df[etf_df['買賣超張數'] > 0].nlargest(10, '買賣超張數')
                            if len(etf_buy_top10) > 0:
                                etf_buy_output = etf_buy_top10[['證券代號', '證券名稱', '買賣超張數']].copy()
                                etf_buy_output['日期'] = file_date
                                etf_buy_output['類別'] = 'ETF買超'
                                etf_buy_output['排名'] = range(1, len(etf_buy_output) + 1)

                                if file_date in stock_daily_prices:
                                    etf_buy_output['收盤價'] = etf_buy_output['證券代號'].apply(
                                        lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('收盤價', '')
                                    )
                                    etf_buy_output['漲跌價差'] = etf_buy_output['證券代號'].apply(
                                        lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('漲跌價差', '')
                                    )
                                else:
                                    etf_buy_output['收盤價'] = ''
                                    etf_buy_output['漲跌價差'] = ''

                                etf_daily_data.append(etf_buy_output)

                            # ETF賣超
                            etf_sell_top10 = etf_df[etf_df['買賣超張數'] < 0].nsmallest(10, '買賣超張數')
                            if len(etf_sell_top10) > 0:
                                etf_sell_output = etf_sell_top10[['證券代號', '證券名稱', '買賣超張數']].copy()
                                etf_sell_output['日期'] = file_date
                                etf_sell_output['類別'] = 'ETF賣超'
                                etf_sell_output['排名'] = range(1, len(etf_sell_output) + 1)

                                if file_date in stock_daily_prices:
                                    etf_sell_output['收盤價'] = etf_sell_output['證券代號'].apply(
                                        lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('收盤價', '')
                                    )
                                    etf_sell_output['漲跌價差'] = etf_sell_output['證券代號'].apply(
                                        lambda x: stock_daily_prices[file_date].get(normalize_stock_code(x), {}).get('漲跌價差', '')
                                    )
                                else:
                                    etf_sell_output['收盤價'] = ''
                                    etf_sell_output['漲跌價差'] = ''

                                etf_daily_data.append(etf_sell_output)

                    df_full = df.copy()
                    df_full['檔案來源'] = os.path.basename(file_path)
                    all_data.append(df_full)

        except Exception as e:
            print(f"讀取檔案 {file_path} 時發生錯誤:{e}")

    statistics = {
        'filtered_out_count': filtered_out_count,
        'processed_count': processed_count
    }

    if allowed_stock_codes is not None:
        print(f"\n{'='*80}")
        print(f"股票代碼過濾統計:")
        print(f"  - 允許的股票代碼總數: {len(allowed_stock_codes)}")
        print(f"  - 處理的股票筆數: {processed_count}")
        print(f"  - 過濾掉的股票筆數: {filtered_out_count}")
        if len(etf_stock_codes) > 0:
            print(f"  - ETF股票數量: {len(etf_stock_codes)}")
        print(f"{'='*80}")

    return (all_data, daily_buy_sell_data, etf_daily_data, buy_top20_tracker,
            sell_top20_tracker, daily_buy_stocks, daily_sell_stocks,
            daily_all_stocks, all_historical_data, statistics)

def organize_daily_buy_sell_data_for_html(daily_buy_sell_data_list):
    """
    將 daily_buy_sell_data 從 DataFrame list 轉換為 HTML 需要的字典格式
    
    Args:
        daily_buy_sell_data_list: list of DataFrames 或 list of dicts
    
    Returns:
        list of dicts: 每個字典包含 {'日期': date, '買超': [...], '賣超': [...]}
    """
    # 如果已經是字典列表，直接返回
    if daily_buy_sell_data_list and isinstance(daily_buy_sell_data_list, list):
        if len(daily_buy_sell_data_list) > 0 and isinstance(daily_buy_sell_data_list[0], dict):
            if '日期' in daily_buy_sell_data_list[0] and '買超' in daily_buy_sell_data_list[0]:
                return daily_buy_sell_data_list
    
    # 按日期分組
    date_data_map = {}
    
    for df in daily_buy_sell_data_list:
        if hasattr(df, 'empty') and df.empty:
            continue
            
        date = df['日期'].iloc[0] if '日期' in df.columns else ''
        category = df['類別'].iloc[0] if '類別' in df.columns else ''
        
        if date not in date_data_map:
            date_data_map[date] = {'日期': date, '買超': [], '賣超': []}
        
        # 轉換 DataFrame 為字典列表
        for _, row in df.iterrows():
            stock_dict = {
                '證券代號': str(row.get('證券代號', '')),
                '證券名稱': str(row.get('證券名稱', '')),
                '買賣超張數': int(row.get('買賣超張數', 0)),
                '收盤價': row.get('收盤價', ''),
                '漲跌': row.get('漲跌價差', '')
            }
            
            # 處理漲跌數值
            price_diff_str = str(stock_dict['漲跌'])
            if price_diff_str and price_diff_str not in ['', '--', 'X', 'nan']:
                try:
                    clean_value = price_diff_str.replace(',', '').replace('+', '')
                    stock_dict['漲跌'] = float(clean_value)
                except:
                    stock_dict['漲跌'] = 0
            else:
                stock_dict['漲跌'] = 0
            
            if category == '買超':
                date_data_map[date]['買超'].append(stock_dict)
            elif category == '賣超':
                date_data_map[date]['賣超'].append(stock_dict)
    
    # 轉換為列表並按日期排序（最新的在前面）
    result = list(date_data_map.values())
    result.sort(key=lambda x: x['日期'], reverse=True)
    
    return result

def organize_daily_buy_sell_data(daily_buy_sell_data_list):
    """
    將 daily_buy_sell_data 從 DataFrame list 轉換為需要的字典格式
    
    Args:
        daily_buy_sell_data_list: list of DataFrames
    
    Returns:
        list of dicts: 每個字典包含 {'日期': date, '買超': [...], '賣超': [...]}
    """
    # 按日期分組
    date_data_map = {}
    
    for df in daily_buy_sell_data_list:
        if df.empty:
            continue
            
        date = df['日期'].iloc[0] if '日期' in df.columns else ''
        category = df['類別'].iloc[0] if '類別' in df.columns else ''
        
        if date not in date_data_map:
            date_data_map[date] = {'日期': date, '買超': [], '賣超': []}
        
        # 轉換 DataFrame 為字典列表
        for _, row in df.iterrows():
            stock_dict = {
                '證券代號': str(row.get('證券代號', '')),
                '證券名稱': str(row.get('證券名稱', '')),
                '買賣超張數': int(row.get('買賣超張數', 0)),
                '收盤價': row.get('收盤價', ''),
                '漲跌': row.get('漲跌價差', '')  # 改名為 '漲跌'
            }
            
            # 處理漲跌數值
            price_diff_str = str(stock_dict['漲跌'])
            if price_diff_str and price_diff_str not in ['', '--', 'X', 'nan']:
                try:
                    # 移除逗號並轉換為數值
                    clean_value = price_diff_str.replace(',', '')
                    stock_dict['漲跌'] = float(clean_value)
                except:
                    stock_dict['漲跌'] = 0
            else:
                stock_dict['漲跌'] = 0
            
            if category == '買超':
                date_data_map[date]['買超'].append(stock_dict)
            elif category == '賣超':
                date_data_map[date]['賣超'].append(stock_dict)
    
    # 轉換為列表並按日期排序（最新的在前面）
    result = list(date_data_map.values())
    result.sort(key=lambda x: x['日期'], reverse=True)
    
    return result

# 【第二步-calculate_stock_statistics】
# 從第二步程式複製 calculate_stock_statistics 函數
def calculate_stock_statistics(all_historical_data, sigma_threshold):
    """
    計算每個證券的統計數據(使用今天往前60天，不含今天)

    Returns:
        dict: {證券代號: {'平均值': x, '標準差': y, '最新值': z, 'Z分數': w, '異常': bool}}
    """
    print(f"\n{'='*80}")
    print("計算過去60天的標準差...")
    print(f"{'='*80}")

    stock_statistics = {}

    for stock_code, date_values in all_historical_data.items():
        if len(date_values) >= 30:
            sorted_values = sorted(date_values, key=lambda x: x[0], reverse=True)
            latest_value = sorted_values[0][1] if len(sorted_values) > 0 else 0
            historical_values = [v[1] for v in sorted_values[1:61]]

            if len(historical_values) >= 30:
                mean = np.mean(historical_values)
                std = np.std(historical_values)

                if std > 0:
                    z_score = abs((latest_value - mean) / std)
                else:
                    z_score = 0

                stock_statistics[stock_code] = {
                    '平均值': mean,
                    '標準差': std,
                    '最新值': latest_value,
                    'Z分數': z_score,
                    '異常': z_score >= sigma_threshold
                }

    return stock_statistics

# 【第二步-analyze_new_entries_and_observables】
# 從第二步程式複製 analyze_new_entries_and_observables 函數
def analyze_new_entries_and_observables(latest_file, daily_buy_stocks, daily_sell_stocks,
                                        daily_all_stocks, stock_statistics, allowed_stock_codes,
                                        sigma_threshold, top_buy_count=50, top_sell_count=20):
    """
    找出最新一天的新進榜證券和值得觀察證券

    Args:
        top_buy_count: 買超分析前N名 (預設50)
        top_sell_count: 賣超分析前N名 (預設20)

    Returns:
        tuple: (new_buy_stocks, new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
                latest_date, latest_buy_stocks_n, latest_sell_stocks_n)
    """
    sorted_dates = sorted(daily_buy_stocks.keys(), reverse=True)
    observable_buy_stocks = {}
    observable_sell_stocks = {}
    new_buy_stocks = set()
    new_sell_stocks = set()
    latest_buy_stocks_n = set()
    latest_sell_stocks_n = set()
    latest_date = None

    if len(sorted_dates) >= 2:
        latest_date = sorted_dates[0]
        previous_dates = sorted_dates[1:]

        latest_df = pd.read_csv(latest_file, encoding='utf-8')

        if '證券代號' in latest_df.columns:
            latest_df['證券代號'] = latest_df['證券代號'].apply(normalize_stock_code)

        if allowed_stock_codes is not None:
            latest_df = latest_df[latest_df['證券代號'].isin(allowed_stock_codes)]

        latest_df['三大法人買賣超股數'] = pd.to_numeric(
            latest_df['三大法人買賣超股數'].astype(str).str.replace(',', ''),
            errors='coerce'
        )
        latest_df['買賣超張數'] = (latest_df['三大法人買賣超股數'] / 1000).fillna(0).astype(int)

        # 使用參數控制的數量
        buy_top_n = latest_df[latest_df['買賣超張數'] > 0].nlargest(top_buy_count, '買賣超張數')
        sell_top_n = latest_df[latest_df['買賣超張數'] < 0].nsmallest(top_sell_count, '買賣超張數')

        latest_buy_stocks_n = set(buy_top_n['證券代號'].tolist())
        latest_sell_stocks_n = set(sell_top_n['證券代號'].tolist())

        # 計算新進榜
        previous_buy_stocks = set()
        previous_sell_stocks = set()
        for date in previous_dates[:4]:
            if date in daily_buy_stocks:
                previous_buy_stocks.update(daily_buy_stocks[date])
            if date in daily_sell_stocks:
                previous_sell_stocks.update(daily_sell_stocks[date])

        latest_buy_stocks = daily_buy_stocks.get(latest_date, set())
        latest_sell_stocks = daily_sell_stocks.get(latest_date, set())

        new_buy_stocks = latest_buy_stocks - previous_buy_stocks
        new_sell_stocks = latest_sell_stocks - previous_sell_stocks

        # 買超值得觀察
        for stock_code in latest_buy_stocks_n:
            reasons = []
            z_score = 0
            mean_val = 0
            std_val = 0

            if stock_code in stock_statistics and stock_statistics[stock_code]['異常']:
                z_score = stock_statistics[stock_code]['Z分數']
                mean_val = stock_statistics[stock_code]['平均值']
                std_val = stock_statistics[stock_code]['標準差']
                reasons.append(f'異常波動({z_score:.1f}σ)')

            positive_days = 0
            for date in previous_dates[:4]:
                if date in daily_all_stocks and stock_code in daily_all_stocks[date]:
                    if daily_all_stocks[date][stock_code] > 0:
                        positive_days += 1
            if positive_days >= 3:
                reasons.append('連續買超')

            if reasons:
                observable_buy_stocks[stock_code] = ('+'.join(reasons), z_score, mean_val, std_val)

        # 賣超值得觀察
        for stock_code in latest_sell_stocks_n:
            reasons = []
            z_score = 0
            mean_val = 0
            std_val = 0

            if stock_code in stock_statistics and stock_statistics[stock_code]['異常']:
                z_score = stock_statistics[stock_code]['Z分數']
                mean_val = stock_statistics[stock_code]['平均值']
                std_val = stock_statistics[stock_code]['標準差']
                reasons.append(f'異常波動({z_score:.1f}σ)')

            negative_days = 0
            for date in previous_dates[:4]:
                if date in daily_all_stocks and stock_code in daily_all_stocks[date]:
                    if daily_all_stocks[date][stock_code] < 0:
                        negative_days += 1
            if negative_days >= 3:
                reasons.append('連續賣超')

            if reasons:
                observable_sell_stocks[stock_code] = ('+'.join(reasons), z_score, mean_val, std_val)

        print(f"\n{'='*80}")
        print(f"【{latest_date} 分析結果】")
        print(f"{'='*80}")
        print(f"使用標準差閾值: {sigma_threshold} 個標準差")
        print(f"買超前20新進榜: {len(new_buy_stocks)} 檔")
        if new_buy_stocks:
            print(f"  證券代號: {', '.join(sorted(new_buy_stocks))}")
        print(f"賣超前20新進榜: {len(new_sell_stocks)} 檔")
        if new_sell_stocks:
            print(f"  證券代號: {', '.join(sorted(new_sell_stocks))}")
        print(f"\n買超前{top_buy_count}值得觀察: {len(observable_buy_stocks)} 檔")
        if observable_buy_stocks:
            for code, (reason, z, mean_val, std_val) in sorted(observable_buy_stocks.items()):
                print(f"  {code}: {reason}")
        print(f"賣超前{top_sell_count}值得觀察: {len(observable_sell_stocks)} 檔")
        if observable_sell_stocks:
            for code, (reason, z, mean_val, std_val) in sorted(observable_sell_stocks.items()):
                print(f"  {code}: {reason}")

    return (new_buy_stocks, new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
            latest_date, latest_buy_stocks_n, latest_sell_stocks_n)


# 【第二步-collect_stock_history】
# 從第二步程式複製 collect_stock_history 函數
def collect_stock_history(latest_buy_stocks_n, latest_sell_stocks_n, folder_path, stock_daily_folder,
                          history_folder, allowed_stock_codes):
    """收集買超前N檔和賣超前N檔股票的歷史數據"""
    print(f"\n{'='*80}")
    print(f"開始收集買超前{len(latest_buy_stocks_n)}檔 + 賣超前{len(latest_sell_stocks_n)}檔股票的歷史數據...")
    print(f"{'='*80}")

    # 合併買超和賣超的股票代碼
    all_target_stocks = latest_buy_stocks_n.union(latest_sell_stocks_n)
    
    if len(all_target_stocks) == 0:
        print(f"沒有股票需要收集歷史數據")
        return

    print(f"總共需要收集 {len(all_target_stocks)} 檔股票的歷史數據")
    print(f"  - 買超: {len(latest_buy_stocks_n)} 檔")
    print(f"  - 賣超: {len(latest_sell_stocks_n)} 檔")
    print(f"  - 重複: {len(latest_buy_stocks_n & latest_sell_stocks_n)} 檔")

    stock_history_data = {}
    for stock_code in all_target_stocks:
        stock_history_data[stock_code] = {}

    # 從 StockTSEShares 讀取
    print("\n從 StockTSEShares 收集數據(2025-01-01 之後)...")
    all_shares_files = glob.glob(os.path.join(folder_path, '*.csv'))

    shares_files_2025 = []
    for file_path in all_shares_files:
        file_date = os.path.basename(file_path).replace('.csv', '')
        if file_date >= '2025-01-01':
            shares_files_2025.append(file_path)

    shares_files_2025 = sorted(shares_files_2025, key=lambda x: os.path.basename(x).replace('.csv', ''), reverse=True)
    print(f"找到 {len(shares_files_2025)} 個 StockTSEShares 檔案(2025-01-01 之後)")

    shares_processed = 0
    for file_path in shares_files_2025:
        try:
            df = pd.read_csv(file_path, encoding='utf-8')

            if '證券代號' in df.columns:
                df['證券代號'] = df['證券代號'].apply(normalize_stock_code)

            if allowed_stock_codes is not None:
                df = df[df['證券代號'].isin(allowed_stock_codes)]

            file_date = os.path.basename(file_path).replace('.csv', '')

            for stock_code in all_target_stocks:
                stock_data = df[df['證券代號'] == stock_code]
                if len(stock_data) > 0:
                    row = stock_data.iloc[0]

                    if file_date not in stock_history_data[stock_code]:
                        stock_history_data[stock_code][file_date] = {
                            '日期': file_date,
                            '股票代碼': stock_code,
                            '股票名稱': row.get('證券名稱', '').strip()
                        }

                    stock_history_data[stock_code][file_date]['外陸資買賣超張數'] = shares_to_lots(row.get('外陸資買賣超股數(不含外資自營商)', 0))
                    stock_history_data[stock_code][file_date]['投信買賣超張數'] = shares_to_lots(row.get('投信買賣超股數', 0))
                    stock_history_data[stock_code][file_date]['自營商買賣超張數'] = shares_to_lots(row.get('自營商買賣超股數', 0))

            shares_processed += 1

        except Exception as e:
            print(f"讀取StockTSEShares檔案 {file_path} 時發生錯誤: {e}")

    print(f"成功處理 {shares_processed} 個 StockTSEShares 檔案")

    # 從 StockTSEDaily 讀取
    if os.path.exists(stock_daily_folder):
        print("\n從 StockTSEDaily 收集數據(2025-01-01 之後)...")

        all_daily_files = glob.glob(os.path.join(stock_daily_folder, '*.csv'))

        daily_files_2025 = []
        for file_path in all_daily_files:
            file_date = os.path.basename(file_path).replace('.csv', '')
            if file_date >= '2025-01-01':
                daily_files_2025.append(file_path)

        daily_files_2025 = sorted(daily_files_2025, key=lambda x: os.path.basename(x).replace('.csv', ''), reverse=True)
        print(f"找到 {len(daily_files_2025)} 個 StockTSEDaily 檔案(2025-01-01 之後)")

        stock_data_count = {code: 0 for code in all_target_stocks}
        daily_processed = 0

        for daily_file in daily_files_2025:
            try:
                # 先嘗試 cp950 編碼，失敗則用 utf-8
                try:
                    df_daily = pd.read_csv(daily_file, encoding='cp950', low_memory=False)
                except:
                    df_daily = pd.read_csv(daily_file, encoding='utf-8', low_memory=False)

                file_date = os.path.basename(daily_file).replace('.csv', '')

                if '證券代號' in df_daily.columns:
                    df_daily['證券代號'] = df_daily['證券代號'].apply(normalize_stock_code)

                if allowed_stock_codes is not None:
                    df_daily = df_daily[df_daily['證券代號'].isin(allowed_stock_codes)]

                for stock_code in all_target_stocks:
                    stock_data = df_daily[df_daily['證券代號'] == stock_code]

                    if len(stock_data) > 0:
                        row = stock_data.iloc[0]

                        if file_date not in stock_history_data[stock_code]:
                            stock_history_data[stock_code][file_date] = {
                                '日期': file_date,
                                '股票代碼': stock_code,
                                '股票名稱': row.get('證券名稱', '').strip()
                            }

                        stock_history_data[stock_code][file_date]['成交張數'] = shares_to_lots(row.get('成交股數', 0))
                        stock_history_data[stock_code][file_date]['成交筆數'] = row.get('成交筆數', '')
                        stock_history_data[stock_code][file_date]['成交金額'] = row.get('成交金額', '')
                        stock_history_data[stock_code][file_date]['開盤價'] = row.get('開盤價', '')
                        stock_history_data[stock_code][file_date]['最高價'] = row.get('最高價', '')
                        stock_history_data[stock_code][file_date]['最低價'] = row.get('最低價', '')
                        stock_history_data[stock_code][file_date]['收盤價'] = row.get('收盤價', '')
                        stock_history_data[stock_code][file_date]['本益比'] = row.get('本益比', '')

                        stock_data_count[stock_code] += 1

                daily_processed += 1

            except Exception as e:
                print(f"讀取StockTSEDaily檔案 {daily_file} 時發生錯誤: {e}")

        print(f"成功處理 {daily_processed} 個 StockTSEDaily 檔案")

        print(f"\n資料統計(前5檔股票):")
        for i, code in enumerate(list(all_target_stocks)[:5]):
            shares_count = len([d for d in stock_history_data[code].keys()])
            daily_count = stock_data_count[code]
            print(f"  {code}: 總共 {shares_count} 天資料,其中 {daily_count} 天有價格資料")
    else:
        print(f"\n警告: StockTSEDaily 資料夾不存在: {stock_daily_folder}")

    # 儲存歷史數據
    print("\n儲存歷史數據到 StockTSEHistory...")

    if not os.path.exists(history_folder):
        os.makedirs(history_folder, exist_ok=True)
        print(f"已建立資料夾: {history_folder}")

    saved_count = 0
    for stock_code, date_dict in stock_history_data.items():
        if len(date_dict) > 0:
            history_list = list(date_dict.values())
            history_df = pd.DataFrame(history_list)

            column_order = [
                '日期', '股票代碼', '股票名稱',
                '成交張數', '成交筆數', '成交金額',
                '開盤價', '最高價', '最低價', '收盤價',
                '本益比', '外陸資買賣超張數', '投信買賣超張數', '自營商買賣超張數'
            ]

            existing_columns = [col for col in column_order if col in history_df.columns]
            history_df = history_df[existing_columns]
            history_df = history_df.sort_values('日期', ascending=False)

            output_file = os.path.join(history_folder, f"{stock_code}.csv")
            history_df.to_csv(output_file, index=False, encoding='utf-8-sig')
            saved_count += 1

            if saved_count <= 5:
                print(f"  已儲存: {stock_code}.csv ({len(history_list)} 筆記錄)")

    print(f"\n完成! 共儲存 {saved_count} 個股票的歷史數據到: {history_folder}")
    print(f"每個檔案包含最近100天的合併數據(StockTSEDaily + StockTSEShares)")
    print(f"注意: 所有股數欄位已轉換為張數(除以1000取整數)")

# 【第二步-aggregate_analysis】
# 從第二步程式複製 aggregate_analysis 函數
def aggregate_analysis(buy_top20_tracker, sell_top20_tracker, stock_sector_map, aggregate_threshold=10000, show_top_n=None):
    """
    彙整分析買超前20和賣超前20

    Args:
        aggregate_threshold: 彙整分析的買賣超張數閾值 (當 show_top_n 為 None 時使用)
        show_top_n: 直接顯示前 N 名 (如果設定此參數，則忽略 aggregate_threshold)

    Returns:
        tuple: (buy_stocks, sell_stocks, both_stocks_set, both_stocks_df)
    """
    if not buy_top20_tracker or not sell_top20_tracker:
        return None, None, set(), None

    all_tracker = buy_top20_tracker + sell_top20_tracker
    all_df = pd.DataFrame(all_tracker)

    summary = all_df.groupby(['證券代號', '證券名稱']).agg({
        '買賣超張數': 'sum',
        '日期': 'count'
    }).reset_index()
    summary.columns = ['證券代號', '證券名稱', '買賣超總和', '出現次數']
    summary['買賣超總和'] = summary['買賣超總和'].astype(int)

    buy_summary = summary[summary['買賣超總和'] > 0].copy()
    buy_summary.columns = ['證券代號', '證券名稱', '買超總和', '買超出現次數']

    sell_summary = summary[summary['買賣超總和'] < 0].copy()
    sell_summary.columns = ['證券代號', '證券名稱', '賣超總和', '賣超出現次數']

    # 找出同時出現在買賣超的證券
    buy_dates_by_stock = {}
    sell_dates_by_stock = {}

    for item in buy_top20_tracker:
        stock_code = item['證券代號']
        if stock_code not in buy_dates_by_stock:
            buy_dates_by_stock[stock_code] = []
        buy_dates_by_stock[stock_code].append(item['日期'])

    for item in sell_top20_tracker:
        stock_code = item['證券代號']
        if stock_code not in sell_dates_by_stock:
            sell_dates_by_stock[stock_code] = []
        sell_dates_by_stock[stock_code].append(item['日期'])

    all_buy_stocks = set(buy_dates_by_stock.keys())
    all_sell_stocks = set(sell_dates_by_stock.keys())
    both_stocks_set = all_buy_stocks & all_sell_stocks

    print(f"\n{'='*80}")
    print(f"發現 {len(both_stocks_set)} 檔證券同時出現在買超前20和賣超前20")
    print("(在5天內,有些天進買超榜、有些天進賣超榜)")
    print(f"{'='*80}")

    # 買超分析
    print(f"\n{'='*80}")
    if show_top_n is not None:
        print(f"【買超分析】最近5天買賣超淨值排名前{show_top_n}名")
    elif aggregate_threshold > 0:
        print(f"【買超分析】最近5天買賣超淨值為正且>={aggregate_threshold}張的證券")
    else:
        print(f"【買超分析】最近5天買賣超淨值為正的所有證券")
    print("(買賣超淨值 = 5天內所有買賣超張數的總和)")
    print(f"{'='*80}\n")

    # 根據參數決定篩選方式
    if show_top_n is not None:
        buy_stocks = buy_summary.sort_values('買超總和', ascending=False).head(show_top_n).copy()
    else:
        buy_stocks = buy_summary[buy_summary['買超總和'] >= aggregate_threshold].sort_values('買超總和', ascending=False).copy()

    buy_stocks['證券領域'] = buy_stocks['證券代號'].apply(lambda x: get_stock_sector(x, stock_sector_map))
    buy_stocks['注意事項'] = buy_stocks['證券代號'].apply(
        lambda x: '⚠️同時出現在賣超' if x in both_stocks_set else ''
    )

    display_buy_stocks = buy_stocks.copy()
    display_buy_stocks['買超總和'] = display_buy_stocks['買超總和'].apply(lambda x: f"{x:,}")

    if len(buy_stocks) > 0:
        print(display_buy_stocks.to_string(index=False))
        print(f"\n共找到 {len(buy_stocks)} 檔符合條件的證券")
    else:
        print("沒有找到符合條件的證券")

    # 賣超分析
    print(f"\n{'='*80}")
    if show_top_n is not None:
        print(f"【賣超分析】最近5天買賣超淨值排名前{show_top_n}名(由大到小)")
    elif aggregate_threshold > 0:
        print(f"【賣超分析】最近5天買賣超淨值為負且<=-{aggregate_threshold}張的證券")
    else:
        print(f"【賣超分析】最近5天買賣超淨值為負的所有證券")
    print("(買賣超淨值 = 5天內所有買賣超張數的總和)")
    print(f"{'='*80}\n")

    # 根據參數決定篩選方式
    if show_top_n is not None:
        sell_stocks = sell_summary.sort_values('賣超總和', ascending=True).head(show_top_n).copy()
    else:
        sell_stocks = sell_summary[sell_summary['賣超總和'] <= -aggregate_threshold].sort_values('賣超總和', ascending=True).copy()

    sell_stocks['證券領域'] = sell_stocks['證券代號'].apply(lambda x: get_stock_sector(x, stock_sector_map))
    sell_stocks['注意事項'] = sell_stocks['證券代號'].apply(
        lambda x: '⚠️同時出現在買超' if x in both_stocks_set else ''
    )

    display_sell_stocks = sell_stocks.copy()
    display_sell_stocks['賣超總和'] = display_sell_stocks['賣超總和'].apply(lambda x: f"{x:,}")

    if len(sell_stocks) > 0:
        print(display_sell_stocks.to_string(index=False))
        print(f"\n共找到 {len(sell_stocks)} 檔符合條件的證券")
    else:
        print("沒有找到符合條件的證券")

    # 同時出現在買賣超的證券詳細分析
    both_stocks_df = None
    if len(both_stocks_set) > 0:
        print(f"\n{'='*80}")
        print("【特別注意】同時出現在買超前20和賣超前20的證券")
        print("(在5天內,有些天進買超榜、有些天進賣超榜)")
        print(f"{'='*80}\n")

        # 取得所有日期並排序（最新在前）
        all_available_dates = sorted(list(set([item['日期'] for item in all_tracker])), reverse=True)
        
        both_stocks_detail = []
        for stock_code in both_stocks_set:
            stock_all_data = all_df[all_df['證券代號'] == stock_code]
            stock_name = stock_all_data.iloc[0]['證券名稱']
            total_sum = int(stock_all_data['買賣超張數'].sum())

            buy_dates = buy_dates_by_stock.get(stock_code, [])
            sell_dates = sell_dates_by_stock.get(stock_code, [])

            buy_dates_short = [format_date_short(d) for d in sorted(buy_dates)]
            sell_dates_short = [format_date_short(d) for d in sorted(sell_dates)]

            buy_dates_str = ', '.join(buy_dates_short)
            sell_dates_str = ', '.join(sell_dates_short)

            buy_sum = int(all_df[(all_df['證券代號'] == stock_code) & (all_df['買賣超張數'] > 0)]['買賣超張數'].sum())
            sell_sum = int(all_df[(all_df['證券代號'] == stock_code) & (all_df['買賣超張數'] < 0)]['買賣超張數'].sum())

            # 建立過去5天的買賣超狀態 (最新在左，確保顯示所有5個日期)
            date_status = []
            for date in all_available_dates[:5]:  # 只取前5個日期
                day_short = format_date_short(date)
                if date in buy_dates:
                    date_status.append(('buy', day_short))
                elif date in sell_dates:
                    date_status.append(('sell', day_short))
                else:
                    date_status.append(('neutral', day_short))
            
            both_stocks_detail.append({
                '證券代號': stock_code,
                '證券名稱': stock_name,
                '證券領域': get_stock_sector(stock_code, stock_sector_map),
                '買超次數': len(buy_dates),
                '買超日期': buy_dates_str,
                '買超總和': buy_sum,
                '賣超次數': len(sell_dates),
                '賣超日期': sell_dates_str,
                '賣超總和': sell_sum,
                '淨買賣超': total_sum,
                '日期狀態': date_status  # 新增：包含 (狀態, 日期) 的列表
            })

        both_stocks_df = pd.DataFrame(both_stocks_detail)
        both_stocks_df = both_stocks_df.sort_values('淨買賣超', ascending=False)

        display_both = both_stocks_df.copy()
        for col in ['買超總和', '賣超總和', '淨買賣超']:
            display_both[col] = display_both[col].apply(lambda x: f"{x:,}")

        print(display_both.to_string(index=False))
        print(f"\n共 {len(both_stocks_df)} 檔證券")

    return buy_stocks, sell_stocks, both_stocks_set, both_stocks_df

# 【第二步-export_to_excel】
# 從第二步程式複製 export_to_excel 函數
def export_to_excel(output_path, buy_stocks, sell_stocks, both_stocks_set, both_stocks_df,
                   daily_buy_sell_data, etf_daily_data, latest_date, new_buy_stocks,
                   new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
                   stock_sector_map, etf_stock_codes):
    """建立並美化 Excel 檔案"""

    if buy_stocks is None and sell_stocks is None:
        print("沒有數據可以輸出到Excel")
        return

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 工作表1: 彙整分析
        startrow = 0

        summary_df = pd.DataFrame([['【彙整買超分析】最近5天買賣超淨值>=10000張 (淨值=5天買賣超總和)']],
                                 columns=[''])
        summary_df.to_excel(writer, sheet_name='彙整分析', index=False, header=False, startrow=startrow)
        startrow += 2

        buy_stocks_output = buy_stocks[['證券代號', '證券領域', '證券名稱', '買超總和', '注意事項']].copy()
        buy_stocks_output.to_excel(writer, sheet_name='彙整分析', index=False, startrow=startrow)
        startrow += len(buy_stocks_output) + 3

        summary_df2 = pd.DataFrame([['【彙整賣超分析】最近5天買賣超淨值<=-10000張 (淨值=5天買賣超總和)']],
                                  columns=[''])
        summary_df2.to_excel(writer, sheet_name='彙整分析', index=False, header=False, startrow=startrow)
        startrow += 2

        sell_stocks_output = sell_stocks[['證券代號', '證券領域', '證券名稱', '賣超總和', '注意事項']].copy()
        sell_stocks_output.to_excel(writer, sheet_name='彙整分析', index=False, startrow=startrow)
        startrow += len(sell_stocks_output) + 3

        if both_stocks_df is not None and len(both_stocks_set) > 0:
            summary_df3 = pd.DataFrame([['【特別注意】同時出現在買超前20和賣超前20的證券(含日期明細)']],
                                      columns=[''])
            summary_df3.to_excel(writer, sheet_name='彙整分析', index=False, header=False, startrow=startrow)
            startrow += 2
            both_stocks_df.to_excel(writer, sheet_name='彙整分析', index=False, startrow=startrow)

        # 工作表2-6: 每日買賣超
        if daily_buy_sell_data:
            # 檢查是否為 DataFrame list
            if daily_buy_sell_data and len(daily_buy_sell_data) > 0:
                if hasattr(daily_buy_sell_data[0], 'empty'):  # 是 DataFrame
                    daily_df = pd.concat(daily_buy_sell_data, ignore_index=True)
                else:  # 是字典，需要轉換
                    all_rows = []
                    for day_data in daily_buy_sell_data:
                        date = day_data['日期']
                        for stock in day_data.get('買超', []):
                            all_rows.append({
                                '日期': date,
                                '類別': '買超',
                                '排名': stock.get('排名', ''),
                                '證券代號': stock['證券代號'],
                                '證券名稱': stock['證券名稱'],
                                '買賣超張數': stock['買賣超張數'],
                                '收盤價': stock['收盤價'],
                                '漲跌價差': stock['漲跌']
                            })
                        for stock in day_data.get('賣超', []):
                            all_rows.append({
                                '日期': date,
                                '類別': '賣超',
                                '排名': stock.get('排名', ''),
                                '證券代號': stock['證券代號'],
                                '證券名稱': stock['證券名稱'],
                                '買賣超張數': stock['買賣超張數'],
                                '收盤價': stock['收盤價'],
                                '漲跌價差': stock['漲跌']
                            })
                    daily_df = pd.DataFrame(all_rows) if all_rows else pd.DataFrame()
            else:
                daily_df = pd.DataFrame()

            for date in sorted(daily_df['日期'].unique(), reverse=True):
                date_data = daily_df[daily_df['日期'] == date]
                sheet_name = date.replace('-', '')[:8]
                startrow = 0
                is_latest = (date == latest_date)

                # 買超部分
                buy_data = date_data[date_data['類別'] == '買超'].copy()
                if len(buy_data) > 0:
                    top_count = len(buy_data)
                    title_df = pd.DataFrame([[f'【{date} 買超 TOP {top_count}】']], columns=[''])
                    title_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                    startrow += 2

                    buy_data['證券領域'] = buy_data['證券代號'].apply(lambda x: get_stock_sector(x, stock_sector_map))

                    if is_latest:
                        buy_data['新進榜'] = buy_data['證券代號'].apply(
                            lambda x: '🔥NEW' if normalize_stock_code(x) in new_buy_stocks else ''
                        )
                        buy_data['值得觀察'] = buy_data['證券代號'].apply(
                            lambda x: f'👀{observable_buy_stocks[normalize_stock_code(x)][0]}' if normalize_stock_code(x) in observable_buy_stocks else ''
                        )
                        buy_data['統計數據(60天)'] = buy_data['證券代號'].apply(
                            lambda x: f'均:{observable_buy_stocks[normalize_stock_code(x)][2]:.0f} 標差:{observable_buy_stocks[normalize_stock_code(x)][3]:.0f}'
                            if normalize_stock_code(x) in observable_buy_stocks and observable_buy_stocks[normalize_stock_code(x)][2] != 0 else ''
                        )
                        buy_data_output = buy_data[['排名', '證券代號', '證券領域', '證券名稱', '收盤價', '漲跌價差', '買賣超張數', '新進榜', '值得觀察', '統計數據(60天)']].copy()
                    else:
                        buy_data_output = buy_data[['排名', '證券代號', '證券領域', '證券名稱', '收盤價', '漲跌價差', '買賣超張數']].copy()

                    buy_data_output.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
                    startrow += len(buy_data_output) + 3

                # 賣超部分
                sell_data = date_data[date_data['類別'] == '賣超'].copy()
                if len(sell_data) > 0:
                    top_count = len(sell_data)
                    title_df2 = pd.DataFrame([[f'【{date} 賣超 TOP {top_count}】']], columns=[''])
                    title_df2.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                    startrow += 2

                    sell_data['證券領域'] = sell_data['證券代號'].apply(lambda x: get_stock_sector(x, stock_sector_map))

                    if is_latest:
                        sell_data['新進榜'] = sell_data['證券代號'].apply(
                            lambda x: '📉NEW' if normalize_stock_code(x) in new_sell_stocks else ''
                        )
                        sell_data['值得觀察'] = sell_data['證券代號'].apply(
                            lambda x: f'👀{observable_sell_stocks[normalize_stock_code(x)][0]}' if normalize_stock_code(x) in observable_sell_stocks else ''
                        )
                        sell_data['統計數據(60天)'] = sell_data['證券代號'].apply(
                            lambda x: f'均:{observable_sell_stocks[normalize_stock_code(x)][2]:.0f} 標差:{observable_sell_stocks[normalize_stock_code(x)][3]:.0f}'
                            if normalize_stock_code(x) in observable_sell_stocks and observable_sell_stocks[normalize_stock_code(x)][2] != 0 else ''
                        )
                        sell_data_output = sell_data[['排名', '證券代號', '證券領域', '證券名稱', '收盤價', '漲跌價差', '買賣超張數', '新進榜', '值得觀察', '統計數據(60天)']].copy()
                    else:
                        sell_data_output = sell_data[['排名', '證券代號', '證券領域', '證券名稱', '收盤價', '漲跌價差', '買賣超張數']].copy()

                    sell_data_output.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
                    startrow += len(sell_data_output) + 3

                # ETF數據
                if len(etf_stock_codes) > 0 and etf_daily_data:
                    etf_df = pd.concat(etf_daily_data, ignore_index=True)
                    etf_date_data = etf_df[etf_df['日期'] == date]

                    if len(etf_date_data) > 0:
                        etf_buy = etf_date_data[etf_date_data['類別'] == 'ETF買超']
                        if len(etf_buy) > 0:
                            title_df3 = pd.DataFrame([[f'【{date} ETF買超 TOP 10】']], columns=[''])
                            title_df3.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                            startrow += 2

                            etf_buy_output = etf_buy[['排名', '證券代號', '證券名稱', '收盤價', '漲跌價差', '買賣超張數']].copy()
                            etf_buy_output.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
                            startrow += len(etf_buy_output) + 3

                        etf_sell = etf_date_data[etf_date_data['類別'] == 'ETF賣超']
                        if len(etf_sell) > 0:
                            title_df4 = pd.DataFrame([[f'【{date} ETF賣超 TOP 10】']], columns=[''])
                            title_df4.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                            startrow += 2

                            etf_sell_output = etf_sell[['排名', '證券代號', '證券名稱', '收盤價', '漲跌價差', '買賣超張數']].copy()
                            etf_sell_output.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

# 【第二步-beautify_excel】
# 從第二步程式複製 beautify_excel 函數
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修正版 HTML 生成函數 - 加強手機響應式支援
"""

def generate_complete_html(output_path, buy_stocks, sell_stocks, both_stocks_set, both_stocks_df,
                          daily_buy_sell_data, etf_daily_data, latest_date, new_buy_stocks,
                          new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
                          stock_sector_map, etf_stock_codes, market_type='TSE'):
    """生成完整的 HTML 分析報告 - 手機優化版"""
    
    market_name = '上市' if market_type == 'TSE' else '上櫃'
    
    # 準備日期標籤
    date_tabs = []
    if daily_buy_sell_data and len(daily_buy_sell_data) > 0:
        for i, day_data in enumerate(daily_buy_sell_data[:5]):
            date = day_data['日期']
            if len(date) == 8:
                formatted_date = f"{date[0:4]}/{date[4:6]}/{date[6:8]}"
            else:
                formatted_date = date
            date_tabs.append((i + 1, formatted_date, day_data))
    
    # HTML 開始 - 加強手機響應式設計
    html_content = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>{market_name}三大法人分析報告</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: "Microsoft JhengHei", "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 2px;
            min-height: 100vh;
            font-size: 15px; /* 基礎字體縮小 */
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}
        
        .tabs {{
            background: white;
            border-radius: 15px;
            padding: 4px 4px 0 4px; /* 縮小間距 */
            margin-bottom: 5px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        
        .tab-buttons {{
            display: flex;
            gap: 2px; /* 縮小間距 */
            flex-wrap: wrap;
            border-bottom: 2px solid #e2e8f0;
            padding-bottom: 10px;
        }}
        
        .tab-button {{
            padding: 2px 4px; /* 縮小按鈕 */
            border: none;
            background: #f7fafc;
            color: #4a5568;
            cursor: pointer;
            border-radius: 8px 8px 0 0;
            font-size: 1.1em; /* 縮小字體 */
            font-weight: 600;
            transition: all 0.3s ease;
            font-family: "Microsoft JhengHei", sans-serif;
        }}
        
        .tab-button:hover {{
            background: #edf2f7;
        }}
        
        .tab-button.active {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        
        .tab-content {{
            display: none;
            padding: 4px 0; /* 縮小間距 */
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        .section {{
            background: white;
            padding: 4px; /* 縮小間距 */
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            margin-bottom: 5px;
        }}
        
        .section-title {{
            font-size: 1.1em; /* 縮小標題 */
            margin-bottom: 4px;
            padding-bottom: 2px;
            border-bottom: 3px solid #667eea;
            color: #2d3748;
        }}
        
        .section-title.buy {{
            border-bottom-color: #48bb78;
        }}
        
        .section-title.sell {{
            border-bottom-color: #f56565;
        }}
        
        .section-title.etf {{
            border-bottom-color: #ed8936;
        }}
        
        .section-title.attention {{
            border-bottom-color: #ecc94b;
        }}
        
        /* 表格容器 - 允許水平滾動 */
        .table-container {{
            width: 100%;
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
            margin-bottom: 4px;
        }}
        
        table {{
            width: 100%;
            min-width: 400px; /* 最小寬度 */
            border-collapse: collapse;
            background: white;
            font-size: 0.95em; /* 縮小表格字體 */
        }}
        
        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        thead.buy {{
            background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
        }}
        
        thead.sell {{
            background: linear-gradient(135deg, #f56565 0%, #e53e3e 100%);
        }}
        
        thead.etf {{
            background: linear-gradient(135deg, #ed8936 0%, #dd6b20 100%);
        }}
        
        thead.attention {{
            background: linear-gradient(135deg, #ecc94b 0%, #d69e2e 100%);
        }}
        
        th {{
            padding: 2px 4px; /* 縮小間距 */
            text-align: left;
            font-weight: 600;
            font-size: 1.1em;
            white-space: nowrap; /* 標題不換行 */
        }}
        
        td {{
            padding: 2px 4px; /* 縮小間距 */
            border-bottom: 1px solid #e2e8f0;
            font-size: 1.1em;
        }}
        
        tr:hover {{
            background-color: #f7fafc;
        }}
        
        .rank {{
            font-weight: bold;
            color: #667eea;
            font-size: 1em;
        }}
        
        .stock-code {{
            font-weight: 600;
            color: #2d3748;
            white-space: nowrap;
            width: 50px; /* 縮小代號欄寬 */
            max-width: 50px;
        }}
        
        .stock-name {{
            font-weight: 600;
            color: #4a5568;
            max-width: 120px; /* 增加名稱寬度 */
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        
        .volume-positive {{
            color: #e53e3e;
            font-weight: 600;
            white-space: nowrap;
        }}
        
        .volume-negative {{
            color: #38a169;
            font-weight: 600;
            white-space: nowrap;
        }}
        
        .price-up {{
            color: #e53e3e;
            font-weight: 600;
            width: 60px; /* 縮小收盤價欄寬 */
            max-width: 60px;
        }}
        
        .price-down {{
            color: #38a169;
            font-weight: 600;
            width: 60px; /* 縮小收盤價欄寬 */
            max-width: 60px;
        }}
        
        .badge {{
            display: inline-block;
            padding: 2px 6px; /* 縮小徽章 */
            border-radius: 10px;
            font-size: 0.75em;
            font-weight: 600;
            margin-left: 3px;
        }}
        
        .badge-new {{
            background-color: #fed7d7;
            color: #c53030;
        }}
        
        .badge-watch {{
            background-color: #fef5e7;
            color: #d69e2e;
        }}
        
        .badge-alert {{
            background-color: #feebc8;
            color: #c05621;
        }}
        
        .footer {{
            background: white;
            padding: 15px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            text-align: center;
            color: #718096;
            font-size: 1.1em;
        }}
        
        /* 手機專用樣式 */
        @media (max-width: 768px) {{
            body {{
                padding: 2px;
                font-size: 13px;
            }}
            
            .section {{
                padding: 2px;
                margin-bottom: 4px;
            }}
            
            .section-title {{
                font-size: 1.1em;
                margin-bottom: 5px;
            }}
            
            table {{
                font-size: 0.75em; /* 手機進一步縮小 */
                min-width: 350px;
            }}
            
            th, td {{
                padding: 1px 2px; /* 手機更緊湊 */
            }}
            
            .tab-button {{
                padding: 3px 6px;
                font-size: 0.85em;
            }}
            
            .stock-name {{
                max-width: 90px; /* 手機縮短名稱 */
            }}
            
            .stock-code {{
                width: 45px;
                max-width: 45px;
            }}
            
            .price-up, .price-down {{
                width: 55px;
                max-width: 55px;
            }}
        }}
        
        /* 極小螢幕 */
        @media (max-width: 480px) {{
            table {{
                font-size: 0.7em;
                min-width: 320px;
            }}
            
            th, td {{
                padding: 1px 1px;
            }}
            
            .stock-name {{
                max-width: 70px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="tabs">
            <div class="tab-buttons">
                <button class="tab-button active" onclick="switchTab(0)">彙整分析</button>"""
    
    # 添加日期標籤按鈕
    for tab_idx, formatted_date, _ in date_tabs:
        html_content += f"""
                <button class="tab-button" onclick="switchTab({tab_idx})">{formatted_date}</button>"""
    
    html_content += """
            </div>

            <div id="tab-0" class="tab-content active">"""
    
    # ========== Tab 0: 彙整分析 ==========
    # 買超分析
    if buy_stocks is not None and len(buy_stocks) > 0:
        html_content += """
                <div class="section">
                    <h2 class="section-title buy">📈 彙整買超分析</h2>
                    <p style="color: #718096; margin-bottom: 5px; font-size: 1.1em;">最近5天買賣超淨值 >= 10000張</p>
                    <div class="table-container">
                        <table>
                            <thead class="buy">
                                <tr>
                                    <th>代號</th>
                                    <th>領域</th>
                                    <th>名稱</th>
                                    <th>買超總和</th>
                                    <th>注意</th>
                                </tr>
                            </thead>
                            <tbody>
"""
        for _, row in buy_stocks.iterrows():
            code = row["證券代號"]
            sector = row.get("證券領域", "")
            name = row["證券名稱"]
            total = row["買超總和"]
            note = row.get("注意事項", "")
            note_html = f'<span class="badge badge-alert">⚠️</span>' if note else ''
            
            html_content += f"""
                                <tr>
                                    <td class="stock-code">{code}</td>
                                    <td>{sector}</td>
                                    <td class="stock-name" title="{name}">{name}</td>
                                    <td class="volume-positive">{total:,}</td>
                                    <td>{note_html}</td>
                                </tr>
"""
        html_content += """
                            </tbody>
                        </table>
                    </div>
                </div>
                """
    
    # 賣超分析
    if sell_stocks is not None and len(sell_stocks) > 0:
        html_content += """
                <div class="section">
                    <h2 class="section-title sell">📉 彙整賣超分析</h2>
                    <p style="color: #718096; margin-bottom: 5px; font-size: 1.1em;">最近5天買賣超淨值 <= -10000張</p>
                    <div class="table-container">
                        <table>
                            <thead class="sell">
                                <tr>
                                    <th>代號</th>
                                    <th>領域</th>
                                    <th>名稱</th>
                                    <th>賣超總和</th>
                                    <th>注意</th>
                                </tr>
                            </thead>
                            <tbody>
"""
        for _, row in sell_stocks.iterrows():
            code = row["證券代號"]
            sector = row.get("證券領域", "")
            name = row["證券名稱"]
            total = row["賣超總和"]
            note = row.get("注意事項", "")
            note_html = f'<span class="badge badge-alert">⚠️</span>' if note else ''
            
            html_content += f"""
                                <tr>
                                    <td class="stock-code">{code}</td>
                                    <td>{sector}</td>
                                    <td class="stock-name" title="{name}">{name}</td>
                                    <td class="volume-negative">{total:,}</td>
                                    <td>{note_html}</td>
                                </tr>
"""
        html_content += """
                            </tbody>
                        </table>
                    </div>
                </div>
                """
    
    # 特別注意
    if both_stocks_df is not None and len(both_stocks_df) > 0:
        html_content += """
                <div class="section">
                    <h2 class="section-title attention">⚠️ 特別注意</h2>
                    <p style="color: #718096; margin-bottom: 5px; font-size: 1.1em;">同時出現在買超與賣超前20</p>
                    <div class="table-container">
                        <table>
                            <thead class="attention">
                                <tr>
                                    <th>代號</th>
                                    <th>名稱</th>
                                    <th>領域</th>
                                    <th>買超和</th>
                                    <th>賣超和</th>
                                    <th>買賣超日期</th>
                                </tr>
                            </thead>
                            <tbody>
"""
        for _, row in both_stocks_df.iterrows():
            code = row["證券代號"]
            name = row["證券名稱"]
            sector = row.get("證券領域", "")
            buy_total = row.get("買超總和", 0)
            sell_total = row.get("賣超總和", 0)
            date_status = row.get("日期狀態", [])
            
            # 生成帶顏色的日期列表
            date_html_parts = []
            for status, day in date_status:
                if status == 'buy':
                    date_html_parts.append(f'<span style="color: #e53e3e; font-weight: 600;">{day}</span>')
                elif status == 'sell':
                    date_html_parts.append(f'<span style="color: #38a169; font-weight: 600;">{day}</span>')
                else:
                    date_html_parts.append(f'<span style="color: #4a5568;">{day}</span>')
            
            dates_display = ', '.join(date_html_parts)
            
            html_content += f"""
                                <tr>
                                    <td class="stock-code">{code}</td>
                                    <td class="stock-name" title="{name}">{name}</td>
                                    <td>{sector}</td>
                                    <td class="volume-positive">{buy_total:,}</td>
                                    <td class="volume-negative">{sell_total:,}</td>
                                    <td style="font-size: 0.9em;">{dates_display}</td>
                                </tr>
"""
        html_content += """
                            </tbody>
                        </table>
                    </div>
                </div>
"""
    
    html_content += """
            </div>
"""
    
    # ========== Tab 1-5: 每日買賣超 ==========
    for tab_idx, formatted_date, day_data in date_tabs:
        html_content += f"""
            <div id="tab-{tab_idx}" class="tab-content">"""
        
        # 買超 TOP
        buy_data = day_data.get('買超', [])
        if buy_data:
            buy_count = len(buy_data)
            html_content += f"""
                <div class="section">
                    <h2 class="section-title buy">📈 買超 TOP {buy_count} ({formatted_date})</h2>
                    <div class="table-container">
                        <table>
                            <thead class="buy">
                                <tr>
                                    <th>排名</th>
                                    <th>代號</th>
                                    <th>名稱</th>
                                    <th>收盤價</th>
                                    <th>漲跌</th>
                                    <th>買賣超</th>
                                </tr>
                            </thead>
                            <tbody>
"""
            for idx, stock in enumerate(buy_data, 1):
                code = stock.get('證券代號', '')
                name = stock.get('證券名稱', '')
                close_price = stock.get('收盤價', 0)
                price_change = stock.get('漲跌', 0)
                volume = stock.get('買賣超張數', 0)
                
                if isinstance(price_change, (int, float)):
                    if price_change > 0:
                        price_class = 'price-up'
                        price_str = f'+{price_change}'
                    elif price_change < 0:
                        price_class = 'price-down'
                        price_str = str(price_change)
                    else:
                        price_class = ''
                        price_str = '0'
                else:
                    price_class = ''
                    price_str = str(price_change)
                
                html_content += f"""
                                <tr>
                                    <td class="rank">{idx}</td>
                                    <td class="stock-code">{code}</td>
                                    <td class="stock-name" title="{name}">{name}</td>
                                    <td>{close_price}</td>
                                    <td class="{price_class}">{price_str}</td>
                                    <td class="volume-positive">{volume:,}</td>
                                </tr>
"""
            html_content += """
                            </tbody>
                        </table>
                    </div>
                </div>
"""
        
        # 賣超 TOP
        sell_data = day_data.get('賣超', [])
        if sell_data:
            sell_count = len(sell_data)
            html_content += f"""
                <div class="section">
                    <h2 class="section-title sell">📉 賣超 TOP {sell_count} ({formatted_date})</h2>
                    <div class="table-container">
                        <table>
                            <thead class="sell">
                                <tr>
                                    <th>排名</th>
                                    <th>代號</th>
                                    <th>名稱</th>
                                    <th>收盤價</th>
                                    <th>漲跌</th>
                                    <th>買賣超</th>
                                </tr>
                            </thead>
                            <tbody>
"""
            for idx, stock in enumerate(sell_data, 1):
                code = stock.get('證券代號', '')
                name = stock.get('證券名稱', '')
                close_price = stock.get('收盤價', 0)
                price_change = stock.get('漲跌', 0)
                volume = stock.get('買賣超張數', 0)
                
                if isinstance(price_change, (int, float)):
                    if price_change > 0:
                        price_class = 'price-up'
                        price_str = f'+{price_change}'
                    elif price_change < 0:
                        price_class = 'price-down'
                        price_str = str(price_change)
                    else:
                        price_class = ''
                        price_str = '0'
                else:
                    price_class = ''
                    price_str = str(price_change)
                
                html_content += f"""
                                <tr>
                                    <td class="rank">{idx}</td>
                                    <td class="stock-code">{code}</td>
                                    <td class="stock-name" title="{name}">{name}</td>
                                    <td>{close_price}</td>
                                    <td class="{price_class}">{price_str}</td>
                                    <td class="volume-negative">{volume:,}</td>
                                </tr>
"""
            html_content += """
                            </tbody>
                        </table>
                    </div>
                </div>
"""
        
        html_content += """
            </div>
"""
    
    # Footer
    from datetime import datetime
    current_time = datetime.now().strftime('%Y-%m-%d')
    
    html_content += f"""
        </div>
        
        <div class="footer">
            <p>資料來源：台灣證券交易所 | 生成時間：{current_time}</p>
        </div>
    </div>
    
    <script>
        function switchTab(tabIndex) {{
            const allContents = document.querySelectorAll('.tab-content');
            allContents.forEach(content => {{
                content.classList.remove('active');
            }});
            
            const allButtons = document.querySelectorAll('.tab-button');
            allButtons.forEach(button => {{
                button.classList.remove('active');
            }});
            
            document.getElementById('tab-' + tabIndex).classList.add('active');
            allButtons[tabIndex].classList.add('active');
        }}
        
        // 禁止雙指縮放
        document.addEventListener('touchstart', function(e) {{
            if (e.touches.length > 1) {{
                e.preventDefault();
            }}
        }}, {{ passive: false }});
        
        // 禁止手勢縮放
        document.addEventListener('gesturestart', function(e) {{
            e.preventDefault();
        }});
    </script>
</body>
</html>"""
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✓ 手機優化版 HTML 已儲存: {output_path}")

def beautify_excel(output_path):
    """美化 Excel 格式"""
    wb = load_workbook(output_path)

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_fill_buy = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    header_fill_sell = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    header_fill_warning = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_fill_observable = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_fill_etf = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_fill_warning = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    title_fill_etf = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
    new_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    red_font = Font(bold=True, color="FF0000", size=11)
    green_font = Font(bold=True, color="00FF00", size=11)

    title_font = Font(bold=True, size=14, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 20

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        price_diff_col_idx = None
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == '漲跌價差':
                    price_diff_col_idx = cell.column
                    break

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '【' in str(cell.value):
                    if 'ETF' in str(cell.value):
                        cell.fill = title_fill_etf
                    elif '特別注意' in str(cell.value):
                        cell.fill = title_fill_warning
                    else:
                        cell.fill = title_fill
                    cell.font = title_font
                    cell.alignment = center_align
                    cell.border = border
                    max_col = ws.max_column
                    ws.merge_cells(start_row=cell.row, start_column=1,
                                  end_row=cell.row, end_column=max_col)
                    for col in range(1, max_col + 1):
                        ws.cell(row=cell.row, column=col).border = border

                elif cell.value in ['證券名稱', '證券代號', '證券領域', '買超總和(張)', '賣超總和(張)',
                                   '排名', '買賣超張數', '注意事項', '淨買賣超(張)', '買超日期', '賣超日期',
                                   '買超次數', '賣超次數', '買超總和', '賣超總和', '淨買賣超',
                                   '新進榜', '值得觀察', '統計數據(60天)', '收盤價', '漲跌價差']:
                    is_buy_section = False
                    is_warning_section = False
                    is_etf_section = False
                    for check_row in range(cell.row, 0, -1):
                        title_cell = ws.cell(row=check_row, column=1).value
                        if title_cell and isinstance(title_cell, str) and '【' in title_cell:
                            if 'ETF' in title_cell:
                                is_etf_section = True
                            elif '特別注意' in title_cell:
                                is_warning_section = True
                            elif '買超' in title_cell and '賣超' not in title_cell:
                                is_buy_section = True
                            break

                    if cell.value == '新進榜':
                        cell.fill = new_fill
                    elif cell.value == '值得觀察':
                        cell.fill = header_fill_observable
                    elif cell.value == '統計數據(60天)':
                        cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                    elif is_etf_section:
                        cell.fill = header_fill_etf
                    elif is_warning_section:
                        cell.fill = header_fill_warning
                    elif is_buy_section:
                        cell.fill = header_fill_buy
                    else:
                        cell.fill = header_fill_sell
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = center_align
                    cell.border = border

                elif cell.value is not None and cell.value != '':
                    if price_diff_col_idx and cell.column == price_diff_col_idx and cell.row > 1:
                        cell_str = str(cell.value).strip()

                        if cell_str and cell_str not in ['', '--', 'X', 'x']:
                            if cell_str.startswith('+'):
                                cell.font = red_font
                            elif cell_str.startswith('-'):
                                cell.font = green_font
                    elif cell.value == '🔥NEW':
                        cell.font = Font(bold=True, color="FF0000", size=11)
                    elif cell.value == '📉NEW':
                        cell.font = Font(bold=True, color="00A86B", size=11)
                    elif isinstance(cell.value, str) and '👀' in str(cell.value):
                        cell.font = Font(bold=True, color="FF8C00", size=10)
                    cell.border = border

    wb.save(output_path)

def run_step2_analysis(base_dir, market_type):
    """執行第二步：分析程式 (GitHub Actions 版本)"""
    print(f"\n{'🔥'*40}")
    print(f"第二步分析：{market_type} ({'上市' if market_type == 'TSE' else '上櫃'})")
    print(f"{'🔥'*40}\n")

    # 設定配置 (使用當前目錄，不使用 Google Drive)
    config = setup_config(market_type=market_type)
    
    # 根據 TOP_STOCKS_ONLY 決定 history_folder 路徑
    if not TOP_STOCKS_ONLY:
        # 使用 local_ 開頭的資料夾
        if market_type == 'TSE':
            config['history_folder'] = os.path.join(base_dir, 'local_StockTSEHistory')
        else:
            config['history_folder'] = os.path.join(base_dir, 'local_StockOTCHistory')
        
        # 確保資料夾存在
        os.makedirs(config['history_folder'], exist_ok=True)
        print(f"📁 History 資料夾: {config['history_folder']}\n")

    # 讀取股票清單
    allowed_stock_codes, stock_sector_map, etf_stock_codes = load_stock_list(config['market_list_path'])

    # 讀取價格資料
    stock_daily_prices = load_stock_daily_prices(config['stock_daily_folder'], allowed_stock_codes)

    # 取得最新檔案
    latest_61_files = get_latest_files(config['folder_path'], num_files=61)

    # 處理三大法人數據
    (all_data, daily_buy_sell_data_raw, etf_daily_data, buy_top20_tracker,
     sell_top20_tracker, daily_buy_stocks, daily_sell_stocks,
     daily_all_stocks, all_historical_data, statistics) = process_shares_files(
        latest_61_files, 
        allowed_stock_codes, 
        stock_daily_prices,
        stock_sector_map, 
        etf_stock_codes,
        top_buy_count=config['top_buy_count'],
        top_sell_count=config['top_sell_count']
    )
    

    # 保留原始 DataFrame list 用於 Excel
    daily_buy_sell_data = daily_buy_sell_data_raw
    
    # 轉換為字典格式用於 HTML
    daily_buy_sell_data_html = organize_daily_buy_sell_data_for_html(daily_buy_sell_data_raw)
    print(f"\n✓ 已整理 {len(daily_buy_sell_data_html)} 天的買賣超數據")

    # 計算標準差
    stock_statistics = calculate_stock_statistics(all_historical_data, config['sigma_threshold'])

    # 分析新進榜與值得觀察
    (new_buy_stocks, new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
     latest_date, latest_buy_stocks_n, latest_sell_stocks_n) = analyze_new_entries_and_observables(
        latest_61_files[0], daily_buy_stocks, daily_sell_stocks,
        daily_all_stocks, stock_statistics, allowed_stock_codes,
        config['sigma_threshold'],
        top_buy_count=config['top_buy_count'],
        top_sell_count=config['top_sell_count']
    )

    # ========== 根據 TOP_STOCKS_ONLY flag 決定要收集歷史的股票 ==========
    if TOP_STOCKS_ONLY:
        # 只收集買超前150 + 賣超前50 的歷史
        print(f"\n{'='*80}")
        print(f"TOP_STOCKS_ONLY = True: 只收集買超前150 + 賣超前50 的歷史數據")
        print(f"{'='*80}")
        collect_buy_stocks = latest_buy_stocks_n
        collect_sell_stocks = latest_sell_stocks_n
    else:
        # 收集所有 CSV 內股票的歷史
        print(f"\n{'='*80}")
        print(f"TOP_STOCKS_ONLY = False: 收集所有 CSV 內股票的歷史數據")
        print(f"{'='*80}")
        
        # 從所有歷史數據中取得所有股票代碼
        all_stocks_in_csv = set(all_historical_data.keys())
        print(f"從 CSV 檔案中找到 {len(all_stocks_in_csv)} 支股票")
        
        collect_buy_stocks = all_stocks_in_csv
        collect_sell_stocks = set()  # 已經包含在 collect_buy_stocks 中

    # 收集歷史數據
    collect_stock_history(collect_buy_stocks, collect_sell_stocks, config['folder_path'],
                      config['stock_daily_folder'], config['history_folder'],
                      allowed_stock_codes)

    # 彙整分析
    buy_stocks, sell_stocks, both_stocks_set, both_stocks_df = aggregate_analysis(
        buy_top20_tracker, sell_top20_tracker, stock_sector_map,
        aggregate_threshold=config.get('aggregate_threshold', 10000),
        show_top_n=config.get('show_top_n', None)
    )

    # 輸出 Excel
    if buy_stocks is not None and sell_stocks is not None:
        export_to_excel(config['output_path'], buy_stocks, sell_stocks, both_stocks_set,
                       both_stocks_df, daily_buy_sell_data, etf_daily_data, latest_date,
                       new_buy_stocks, new_sell_stocks, observable_buy_stocks,
                       observable_sell_stocks, stock_sector_map, etf_stock_codes)

        # 美化 Excel
        beautify_excel(config['output_path'])

        # 生成 HTML 報告 - 使用轉換後的字典格式
        html_output_path = config['output_path'].replace('.xlsx', '_complete.html')
        
        print(f"\n準備生成 HTML: {html_output_path}")
        print(f"  - buy_stocks: {len(buy_stocks) if buy_stocks is not None else 0} rows")
        print(f"  - sell_stocks: {len(sell_stocks) if sell_stocks is not None else 0} rows")
        print(f"  - daily_buy_sell_data_html: {len(daily_buy_sell_data_html)} days")
        
        generate_complete_html(
            html_output_path, buy_stocks, sell_stocks, both_stocks_set,
            both_stocks_df, daily_buy_sell_data_html, etf_daily_data, latest_date,
            new_buy_stocks, new_sell_stocks, observable_buy_stocks,
            observable_sell_stocks, stock_sector_map, etf_stock_codes,
            market_type=market_type)

        print(f"\n✓ {market_type} 分析完成")
        print(f"✓ Excel 已儲存: {config['output_path']}")
        print(f"✓ HTML 已儲存: {html_output_path}")

    # ========== 儲存買超排名順序 ==========
    # ========== 儲存買超+賣超排名順序 ==========
    if latest_date and latest_buy_stocks_n:
        try:
            # 讀取最新一天的資料來取得完整排名
            latest_file = latest_61_files[0]
            latest_df = pd.read_csv(latest_file, encoding='utf-8')

            if '證券代號' in latest_df.columns:
                latest_df['證券代號'] = latest_df['證券代號'].apply(normalize_stock_code)

            if allowed_stock_codes is not None:
                latest_df = latest_df[latest_df['證券代號'].isin(allowed_stock_codes)]

            latest_df['三大法人買賣超股數'] = pd.to_numeric(
                latest_df['三大法人買賣超股數'].astype(str).str.replace(',', ''),
                errors='coerce'
            )
            latest_df['買賣超張數'] = (latest_df['三大法人買賣超股數'] / 1000).fillna(0).astype(int)

            # 取得買超前N和賣超前N的排名順序
            top_buy_count = config.get('top_buy_count', 50)
            top_sell_count = config.get('top_sell_count', 20)
            
            buy_top = latest_df[latest_df['買賣超張數'] > 0].nlargest(top_buy_count, '買賣超張數')
            sell_top = latest_df[latest_df['買賣超張數'] < 0].nsmallest(top_sell_count, '買賣超張數')
            
            buy_ranking = buy_top['證券代號'].tolist()
            sell_ranking = sell_top['證券代號'].tolist()

            # 儲存排名到檔案（買超+賣超）
            ranking_file = os.path.join(config['output_folder'], f'{market_type}_buy_ranking.txt')
            with open(ranking_file, 'w', encoding='utf-8') as f:
                f.write(f"# {market_type} - {latest_date}\n")
                
                # 寫入買超前N名
                for rank, code in enumerate(buy_ranking, 1):
                    stock_name = latest_df[latest_df['證券代號'] == code]['證券名稱'].iloc[0] if len(latest_df[latest_df['證券代號'] == code]) > 0 else ''
                    buy_amount = latest_df[latest_df['證券代號'] == code]['買賣超張數'].iloc[0] if len(latest_df[latest_df['證券代號'] == code]) > 0 else 0
                    f.write(f"{rank},{code},{stock_name},{buy_amount}\n")
                
                # 寫入賣超前N名
                for rank, code in enumerate(sell_ranking, top_buy_count + 1):
                    stock_name = latest_df[latest_df['證券代號'] == code]['證券名稱'].iloc[0] if len(latest_df[latest_df['證券代號'] == code]) > 0 else ''
                    sell_amount = latest_df[latest_df['證券代號'] == code]['買賣超張數'].iloc[0] if len(latest_df[latest_df['證券代號'] == code]) > 0 else 0
                    f.write(f"{rank},{code},{stock_name},{sell_amount}\n")

            print(f"\n✓ 排名已儲存: {ranking_file}")
            print(f"  買超前{top_buy_count}名 + 賣超前{top_sell_count}名 = 共{top_buy_count + top_sell_count}筆")
            print(f"  買超前10名: {', '.join(buy_ranking[:10])}")
            print(f"  賣超前5名: {', '.join(sell_ranking[:5])}")
        except Exception as e:
            print(f"\n⚠ 儲存排名時發生錯誤: {e}")


# ============================================================================
# 第三步：圖表生成的所有類別和函數
# ============================================================================

# ============================================================================
# 模組 1: 配置管理 (Config)
# ============================================================================

class Config:
    """配置管理類別"""

    # ========== 全域設定 ==========
    OVERWRITE_EXISTING = True  # True: 覆蓋已存在的檔案, False: 跳過已存在的檔案
    MARKET_TYPE = 'TSE'  # 'TSE': 上市, 'OTC': 上櫃, 'ALL': 全部
    RUN_ALL = True  # True: 批次處理所有股票, False: 手動輸入單一股票
    # 批次處理會同時生成: (1) 個別HTML到StockTSEHTML, (2) 合併HTML到StockInfo
    # ==============================

    FONT_PATH = None  # 中文字體路徑

    @staticmethod
    def setup_config(market_type='TSE', base_path='.'):
        """
        設定所有路徑變數

        Args:
            market_type: 'TSE' (上市) 或 'OTC' (上櫃)
            base_path: 基礎路徑 (預設為當前目錄)

        Returns:
            dict: 包含所有路徑配置的字典
        """

        if market_type == 'TSE':
            config = {
                'market_type': market_type,
                'market_name': '上市',
                'history_folder': os.path.join(base_path, 'StockTSEHistory'),
                'html_output_folder': os.path.join(base_path, 'StockTSEHTML'),
                'merged_output_folder': os.path.join(base_path, 'StockInfo'),
                'stocklist_folder': os.path.join(base_path, 'StockList'),
            }
        else:  # OTC
            config = {
                'market_type': market_type,
                'market_name': '上櫃',
                'history_folder': os.path.join(base_path, 'StockOTCHistory'),
                'html_output_folder': os.path.join(base_path, 'StockOTCHTML'),
                'merged_output_folder': os.path.join(base_path, 'StockInfo'),
                'stocklist_folder': os.path.join(base_path, 'StockList'),
            }

        # 建立輸出資料夾
        os.makedirs(config['html_output_folder'], exist_ok=True)
        os.makedirs(config['merged_output_folder'], exist_ok=True)

        print(f"{'='*80}")
        print(f"市場類型: {market_type} ({config['market_name']})")
        print(f"輸出模式: 個別HTML + 合併HTML")
        print(f"歷史數據資料夾: {config['history_folder']}")
        print(f"個別HTML輸出: {config['html_output_folder']}")
        print(f"合併HTML輸出: {config['merged_output_folder']}")
        print(f"{'='*80}\n")

        return config


# ============================================================================
# 模組 2: 工具函數 (Utils)
# ============================================================================

class Utils:
    """工具函數類別"""

    @staticmethod
    def setup_chinese_font(base_path='.'):
        """設定中文字體"""
        font_path = os.path.join(base_path, 'StockList', 'Font.ttf')

        if os.path.exists(font_path):
            Config.FONT_PATH = font_path
            print(f"✓ 找到字體檔案: {font_path}")
        else:
            print(f"⚠ 找不到字體檔案: {font_path}")
            print("  HTML 圖表將使用預設字體")
            Config.FONT_PATH = None

        return Config.FONT_PATH

    @staticmethod
    def read_csv_auto_encoding(file_path):
        """自動偵測編碼讀取 CSV"""
        encodings = ['utf-8-sig', 'utf-8', 'big5', 'cp950']
        for encoding in encodings:
            try:
                return pd.read_csv(file_path, encoding=encoding)
            except:
                continue
        raise ValueError(f"無法讀取檔案: {file_path}")

    @staticmethod
    def get_stock_name(base_path, stock_code):
        """從 StockList 取得股票名稱"""
        try:
            stocklist_path = os.path.join(base_path, 'StockList', 'StockList_simplified.csv')
            if not os.path.exists(stocklist_path):
                return ''

            df = Utils.read_csv_auto_encoding(stocklist_path)

            for code_col in df.columns:
                if '代' in code_col or 'code' in code_col.lower():
                    for name_col in df.columns:
                        if '名' in name_col or 'name' in name_col.lower():
                            matched = df[df[code_col].astype(str) == str(stock_code)]
                            if len(matched) > 0:
                                return str(matched.iloc[0][name_col])
            return ''
        except:
            return ''

    @staticmethod
    def get_all_stock_codes_from_history(history_folder):
        """從 History 資料夾取得所有股票代碼"""
        try:
            if not os.path.exists(history_folder):
                print(f"❌ 找不到資料夾: {history_folder}")
                return []

            csv_files = glob.glob(os.path.join(history_folder, "*.csv"))

            if not csv_files:
                print(f"❌ 資料夾中沒有 CSV 檔案: {history_folder}")
                return []

            stock_codes = []
            for csv_file in csv_files:
                filename = os.path.basename(csv_file)
                stock_code = os.path.splitext(filename)[0]
                stock_codes.append(stock_code)

            stock_codes.sort()

            print(f"✓ 從 {os.path.basename(history_folder)} 找到 {len(stock_codes)} 支股票")
            return stock_codes

        except Exception as e:
            print(f"❌ 讀取資料夾失敗: {str(e)}")
            return []

    @staticmethod
    def prepare_chart_data(df):
        """準備圖表數據"""
        df_chart = df.copy()

        # 確保日期是 datetime 格式
        df_chart['日期'] = pd.to_datetime(df_chart['日期'], errors='coerce')

        # 移除日期為 NaT 的資料
        df_chart = df_chart[df_chart['日期'].notna()]

        df_chart = df_chart.sort_values('日期')
        df_chart = df_chart.tail(60).copy()

        # 確保數值欄位是數字類型
        numeric_cols = ['開盤價', '最高價', '最低價', '收盤價', '成交張數',
                        '外陸資買賣超張數', '投信買賣超張數', '自營商買賣超張數']
        for col in numeric_cols:
            if col in df_chart.columns:
                if df_chart[col].dtype == 'object':
                    df_chart[col] = df_chart[col].astype(str).str.replace(',', '').str.replace('--', '0')
                df_chart[col] = pd.to_numeric(df_chart[col], errors='coerce')

        # 計算 MA5 和 MA10（移動平均線）
        if '收盤價' in df_chart.columns:
            df_chart['MA5'] = df_chart['收盤價'].rolling(window=5, min_periods=1).mean()
            df_chart['MA10'] = df_chart['收盤價'].rolling(window=10, min_periods=1).mean()

        return df_chart


# ============================================================================
# 模組 3: Plotly 圖表生成 (ChartPlotly)
# ============================================================================

class ChartPlotly:
    """Plotly 圖表生成類別"""

    @staticmethod
    def generate_chart(df, stock_code, stock_name, html_output_path=None):
        """
        使用 Plotly 生成互動式技術分析圖表 (HTML)

        Args:
            html_output_path: 如果為 None, 則只返回 HTML 字串不儲存檔案
        """

        df_chart = Utils.prepare_chart_data(df)

        print(f"  圖表數據範圍: {df_chart['日期'].min().strftime('%Y-%m-%d')} ~ {df_chart['日期'].max().strftime('%Y-%m-%d')} (共 {len(df_chart)} 筆)")

        latest_date_str = df_chart['日期'].max().strftime('%Y-%m-%d')

        # 計算統計數據
        stats = ChartPlotly._calculate_statistics(df_chart)

        # 創建子圖（4層：K線、成交量、當日買賣超、累積買賣超）
        fig = make_subplots(
            rows=4, cols=1,
            shared_xaxes=True,
            vertical_spacing=0.03,
            subplot_titles=(
                '',  # 第一層標題留空
                '',  # 第二層標題留空
                '',  # 第三層標題留空
                ''   # 第四層標題留空
            ),
            row_heights=[0.35, 0.15, 0.25, 0.25],
            specs=[[{"secondary_y": False}],
                   [{"secondary_y": False}],
                   [{"secondary_y": False}],
                   [{"secondary_y": False}]]
        )

        # 第一層: K線圖
        ChartPlotly._add_candlestick(fig, df_chart)
        
        # 第一層: 移動平均線 (MA5 和 MA10)
        ChartPlotly._add_moving_averages(fig, df_chart)

        # 第二層: 成交量
        ChartPlotly._add_volume_traces(fig, df_chart)

        # 第三層: 三大法人當日買賣超
        has_institutional = ChartPlotly._add_institutional_daily(fig, df_chart)

        # 第四層: 三大法人累積買賣超
        if has_institutional:
            ChartPlotly._add_institutional_cumulative(fig, df_chart)

        # 更新佈局
        ChartPlotly._update_layout(fig, stock_code, stock_name, latest_date_str, df_chart, stats)

        # 生成 HTML 字串
        html_string = fig.to_html(include_plotlyjs='cdn', div_id=f'chart_{stock_code}')

        # 如果指定了輸出路徑,則儲存完整的 HTML 檔案
        if html_output_path:
            full_html = ChartPlotly._wrap_html(html_string, f"{stock_code} {stock_name}")
            with open(html_output_path, 'w', encoding='utf-8') as f:
                f.write(full_html)
            print(f"  ✓ HTML圖表已儲存: {html_output_path}")

        return html_string

    @staticmethod
    def _wrap_html(chart_html, title="股票圖表"):
        """包裝完整的 HTML 結構"""
        viewport_meta = '<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, minimum-scale=1.0, user-scalable=no">'

        touch_action_css = '''
    <style>
        html {
            -webkit-text-size-adjust: 100%;
            -ms-text-size-adjust: 100%;
        }

        body {
            margin: 0;
            padding: 0;
            overflow-y: auto;
            overflow-x: hidden;
            -webkit-overflow-scrolling: touch;
            -webkit-user-select: none;
            -moz-user-select: none;
            -ms-user-select: none;
            user-select: none;
        }

        .plotly {
            touch-action: pan-y;
            -ms-touch-action: pan-y;
        }

        * {
            -webkit-tap-highlight-color: transparent;
        }

        .stock-separator {
            height: 30px;
            background: linear-gradient(to bottom, #f0f0f0, #ffffff);
            margin: 20px 0;
            border-top: 2px solid #ddd;
            border-bottom: 2px solid #ddd;
        }
    </style>'''

        disable_gestures_script = '''
    <script>
        document.addEventListener('DOMContentLoaded', function() {
            // 禁止雙指縮放
            document.addEventListener('touchstart', function(e) {
                if (e.touches.length > 1) {
                    e.preventDefault();
                }
            }, { passive: false });

            // 禁止手勢縮放
            document.addEventListener('gesturestart', function(e) {
                e.preventDefault();
            });

            document.addEventListener('gesturechange', function(e) {
                e.preventDefault();
            });

            document.addEventListener('gestureend', function(e) {
                e.preventDefault();
            });

            // 禁止雙擊縮放
            let lastTouchEnd = 0;
            document.addEventListener('touchend', function(e) {
                const now = Date.now();
                if (now - lastTouchEnd <= 300) {
                    e.preventDefault();
                }
                lastTouchEnd = now;
            }, false);

            // 禁止滾輪縮放(Ctrl+滾輪)
            document.addEventListener('wheel', function(e) {
                if (e.ctrlKey) {
                    e.preventDefault();
                }
            }, { passive: false });

            // 禁止橫向滾動
            document.addEventListener('touchmove', function(e) {
                if (!e.target.closest('.plotly')) {
                    const touch = e.touches[0];
                    const deltaX = Math.abs(touch.clientX - (touch.startX || touch.clientX));
                    const deltaY = Math.abs(touch.clientY - (touch.startY || touch.clientY));

                    if (deltaX > deltaY) {
                        e.preventDefault();
                    }
                }
            }, { passive: false });

            document.addEventListener('touchstart', function(e) {
                const touch = e.touches[0];
                touch.startX = touch.clientX;
                touch.startY = touch.clientY;
            }, { passive: true });
        });
    </script>'''

        full_html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    {viewport_meta}
    <title>{title}</title>
    {touch_action_css}
</head>
<body>
{chart_html}
{disable_gestures_script}
</body>
</html>'''

        return full_html

    @staticmethod
    def _calculate_statistics(df_chart):
        """計算統計數據"""
        latest = df_chart.iloc[-1]

        stats = {
            '成交量': latest['成交張數'] if '成交張數' in latest and pd.notna(latest['成交張數']) else 0,
        }

        # 計算法人累積
        if '外陸資買賣超張數' in df_chart.columns:
            foreign_cumsum = df_chart['外陸資買賣超張數'].fillna(0).cumsum()
            stats['外資累積'] = foreign_cumsum.iloc[-1] if len(foreign_cumsum) > 0 else 0
        else:
            stats['外資累積'] = 0

        if '投信買賣超張數' in df_chart.columns:
            trust_cumsum = df_chart['投信買賣超張數'].fillna(0).cumsum()
            stats['投信累積'] = trust_cumsum.iloc[-1] if len(trust_cumsum) > 0 else 0
        else:
            stats['投信累積'] = 0

        if '自營商買賣超張數' in df_chart.columns:
            dealer_cumsum = df_chart['自營商買賣超張數'].fillna(0).cumsum()
            stats['自營累積'] = dealer_cumsum.iloc[-1] if len(dealer_cumsum) > 0 else 0
        else:
            stats['自營累積'] = 0

        return stats

    @staticmethod
    def _add_candlestick(fig, df_chart):
        """新增 K 線圖"""
        fig.add_trace(
            go.Candlestick(
                x=df_chart['日期'],
                open=df_chart['開盤價'],
                high=df_chart['最高價'],
                low=df_chart['最低價'],
                close=df_chart['收盤價'],
                name='K線',
                increasing_line_color='#FF5252',  # 玩股網風格的紅色
                increasing_fillcolor='#FF5252',
                decreasing_line_color='#00C851',  # 玩股網風格的綠色
                decreasing_fillcolor='#00C851',
                line=dict(width=0.8),  # 影線加粗
                xhoverformat="%m-%d",
                yhoverformat=".2f"
            ),
            row=1, col=1
        )

    @staticmethod
    def _add_moving_averages(fig, df_chart):
        """新增移動平均線（只加 MA5 和 MA10）"""
        # 只加 MA5 和 MA10，並檢查欄位是否存在
        for ma_name, ma_col, color in [('MA5', 'MA5', 'blue'),
                                         ('MA10', 'MA10', 'orange')]:
            if ma_col in df_chart.columns:
                # 確保資料不全是 NaN
                if df_chart[ma_col].notna().sum() > 0:
                    fig.add_trace(
                        go.Scatter(
                            x=df_chart['日期'],
                            y=df_chart[ma_col],
                            name=ma_name,
                            line=dict(color=color, width=1.5),
                            mode='lines',
                            hovertemplate=f'{ma_name}: %{{y:.2f}}<extra></extra>'
                        ),
                        row=1, col=1
                    )

    @staticmethod
    def _add_volume_traces(fig, df_chart):
        """新增成交量圖表（美化長條圖樣式）"""
        if '成交張數' in df_chart.columns:
            volume_lots = pd.to_numeric(df_chart['成交張數'], errors='coerce')
            
            # 根據漲跌決定顏色（紅漲綠跌）
            colors = []
            for i in range(len(df_chart)):
                if i == 0:
                    # 第一天用開盤收盤比較
                    if df_chart['收盤價'].iloc[i] >= df_chart['開盤價'].iloc[i]:
                        colors.append('rgba(255, 82, 82, 0.8)')  # 玩股網風格紅色
                    else:
                        colors.append('rgba(0, 200, 81, 0.8)')   # 玩股網風格綠色
                else:
                    # 其他天與前一天收盤價比較
                    if df_chart['收盤價'].iloc[i] >= df_chart['收盤價'].iloc[i-1]:
                        colors.append('rgba(255, 82, 82, 0.8)')  # 玩股網風格紅色
                    else:
                        colors.append('rgba(0, 200, 81, 0.8)')   # 玩股網風格綠色
            
            # 成交量長條圖
            fig.add_trace(
                go.Bar(
                    x=df_chart['日期'],
                    y=volume_lots,
                    name='成交量',
                    marker=dict(
                        color=colors,
                        line=dict(width=0)  # 無邊框更簡潔
                    ),
                    hovertemplate='成交量: %{y:,.0f}張<extra></extra>',
                    showlegend=True
                ),
                row=2, col=1
            )

    @staticmethod
    def _add_institutional_daily(fig, df_chart):
        """新增三大法人當日買賣超"""
        has_institutional_data = False
        if '外陸資買賣超張數' in df_chart.columns:
            foreign = pd.to_numeric(df_chart['外陸資買賣超張數'], errors='coerce')
            trust = pd.to_numeric(df_chart.get('投信買賣超張數', 0), errors='coerce')
            dealer = pd.to_numeric(df_chart.get('自營商買賣超張數', 0), errors='coerce')

            if foreign.notna().sum() > 0 or trust.notna().sum() > 0 or dealer.notna().sum() > 0:
                has_institutional_data = True

                # 統一顏色配置與圖例名稱
                for name, data, color in [
                    ('外資', foreign, 'rgba(255, 82, 82, 0.75)'),    # 鮮明紅色
                    ('投信', trust, 'rgba(0, 200, 81, 0.75)'),       # 鮮明綠色
                    ('自營商', dealer, 'rgba(0, 191, 255, 0.75)')    # 天空藍
                ]:
                    fig.add_trace(
                        go.Bar(
                            x=df_chart['日期'],
                            y=data,
                            name=name,  # 圖例顯示: 外資/投信/自營商
                            marker_color=color,
                            hovertemplate=f'{name}: %{{y:,.0f}}張<extra></extra>',
                            legendgroup=name,  # 將上下圖表的同類型分組
                            showlegend=True
                        ),
                        row=3, col=1
                    )

        return has_institutional_data

    @staticmethod
    def _add_institutional_cumulative(fig, df_chart):
        """新增三大法人累積買賣超（平滑曲線）"""
        if '外陸資買賣超張數' in df_chart.columns:
            foreign_cumsum = pd.to_numeric(df_chart['外陸資買賣超張數'], errors='coerce').fillna(0).cumsum()
            trust_cumsum = pd.to_numeric(df_chart.get('投信買賣超張數', 0), errors='coerce').fillna(0).cumsum()
            dealer_cumsum = pd.to_numeric(df_chart.get('自營商買賣超張數', 0), errors='coerce').fillna(0).cumsum()

            # 統一顏色配置與圖例名稱（使用spline平滑曲線）
            for name, data, color in [
                ('外資', foreign_cumsum, 'rgb(255, 82, 82)'),    # 鮮明紅色
                ('投信', trust_cumsum, 'rgb(0, 200, 81)'),       # 鮮明綠色
                ('自營商', dealer_cumsum, 'rgb(0, 191, 255)')    # 天空藍
            ]:
                fig.add_trace(
                    go.Scatter(
                        x=df_chart['日期'],
                        y=data,
                        name=f'{name}累積',  # 圖例顯示: 外資累積/投信累積/自營商累積
                        line=dict(color=color, width=2.5, shape='spline', smoothing=0.8),
                        mode='lines',
                        hovertemplate=f'{name}累積: %{{y:,.0f}}張<extra></extra>',
                        legendgroup=name,  # 與上層的外資/投信/自營商同組
                        showlegend=True
                    ),
                    row=4, col=1
                )

    @staticmethod
    def _update_layout(fig, stock_code, stock_name, latest_date_str, df_chart, stats):
        """更新圖表佈局"""
        # 建立統計資訊文字 (簡化版，移除MA)
        stats_line1 = (
            f"最新資料日期: {latest_date_str} | "
            f"外資累積: {stats['外資累積']:,.0f}張 | "
            f"投信累積: {stats['投信累積']:,.0f}張 | "
            f"自營累積: {stats['自營累積']:,.0f}張"
        )
        stats_line2 = (
            f"股價K線圖 | "
            f"成交量: {stats['成交量']:,.0f}張"
        )

        fig.update_layout(
            title=dict(
                text=f'{stock_code} {stock_name} 技術分析圖表 (最近60筆)<br><sub>{stats_line1}</sub><br><sub>{stats_line2}</sub>',
                x=0.5,
                xanchor='center',
                font=dict(size=16, family='Microsoft JhengHei, Arial, sans-serif')
            ),
            xaxis_rangeslider_visible=False,
            height=1500,  # 4層圖表高度
            showlegend=True,
            hovermode='x unified',
            template='plotly_white',
            barmode='relative',
            legend=dict(
                orientation="v",
                yanchor="top",
                y=0.98,
                xanchor="left",
                x=0.01,
                bgcolor="rgba(255, 255, 255, 0.8)",
                bordercolor="lightgray",
                borderwidth=1,
                font=dict(family='Microsoft JhengHei, Arial, sans-serif')
            ),
            font=dict(family='Microsoft JhengHei, Arial, sans-serif'),  # 全域字體設定
            dragmode='pan'  # 允許拖曳,但由 fixedrange 限制軸範圍
        )

        # 移除所有子圖標題（已在 make_subplots 中設為空字串）
        # 不需要額外的 annotations 設定

        # 計算股價範圍（只使用OHLC）
        price_cols = ['開盤價', '最高價', '最低價', '收盤價']
        price_min = df_chart[price_cols].min().min()
        price_max = df_chart[price_cols].max().max()
        price_margin = (price_max - price_min) * 0.05
        price_range = [price_min - price_margin, price_max + price_margin]

        # 更新Y軸 - 禁用縮放
        fig.update_yaxes(title_text="股價 (元)", row=1, col=1, range=price_range, fixedrange=True)
        fig.update_yaxes(title_text="成交量 (張)", row=2, col=1, tickformat=",", fixedrange=True)
        fig.update_yaxes(title_text="當日買賣超 (張)", row=3, col=1, tickformat=",", fixedrange=True)
        fig.update_yaxes(title_text="累積買賣超 (張)", row=4, col=1, tickformat=",", fixedrange=True)

        # 更新X軸 - 禁用縮放
        # 更新X軸 - 移除非交易日空隙，讓 K 線顯示完整
        start_date = df_chart['日期'].min()
        end_date = df_chart['日期'].max()
        
        # 獲取實際交易日期列表
        trading_dates = df_chart['日期'].tolist()

        tickvals = []
        current = start_date.replace(day=1)
        while current <= end_date:
            for day in [1, 6, 11, 16, 21, 26]:
                try:
                    tick_date = current.replace(day=day)
                    if start_date <= tick_date <= end_date:
                        tickvals.append(tick_date)
                except:
                    pass
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        for i in range(1, 5):
            fig.update_xaxes(
                tickformat="%m-%d",
                tickangle=-45,
                tickmode='array',
                tickvals=tickvals,
                showticklabels=True,
                autorange=True,  # 自動調整範圍以顯示完整資料
                hoverformat="%m-%d",
                fixedrange=True,  # 禁用 X 軸縮放
                rangebreaks=[
                    dict(values=pd.date_range(start=start_date, end=end_date, freq='D')
                         .difference(pd.DatetimeIndex(trading_dates)).tolist())  # 移除所有非交易日
                ],
                row=i, col=1
            )


# ============================================================================
# 模組 4: 股票處理器 (Processor)
# ============================================================================

class Processor:
    """股票處理類別"""

    @staticmethod
    def process_stock(stock_code, base_path, config, save_individual=True):
        """
        處理單一股票

        Args:
            save_individual: True 則儲存個別檔案, False 則只返回 HTML 字串
        
        Returns:
            HTML 字串 (用於合併), 或 True/False (儲存狀態)
        """

        print(f"\n{'='*70}")
        print(f"處理股票: {stock_code}")
        print('='*70)

        csv_file = os.path.join(config['history_folder'], f"{stock_code}.csv")

        if not os.path.exists(csv_file):
            print(f"❌ 找不到檔案: {csv_file}")
            return None

        print(f"⏳ 讀取 {os.path.basename(config['history_folder'])}/{stock_code}.csv...")

        try:
            result = Utils.read_csv_auto_encoding(csv_file)
            print(f"✓ 成功讀取 {len(result)} 筆資料")
        except Exception as e:
            print(f"❌ 讀取失敗: {str(e)}")
            return None

        stock_name = result['股票名稱'].iloc[0] if '股票名稱' in result.columns and len(result) > 0 else ''
        if not stock_name:
            stock_name = Utils.get_stock_name(base_path, stock_code)

        print(f"✅ 資料載入完成")
        print(f"  股票: {stock_code} {stock_name}")
        print(f"  筆數: {len(result)}")
        if '日期' in result.columns:
            print(f"  日期範圍: {result['日期'].min()} ~ {result['日期'].max()}")

        print(f"⏳ 生成技術分析圖表...")

        try:
            # 生成 HTML 字串 (用於合併)
            html_string = ChartPlotly.generate_chart(
                result,
                stock_code,
                stock_name,
                html_output_path=None
            )
            
            # 如果需要,同時儲存個別檔案
            if save_individual:
                html_output_file = os.path.join(config['html_output_folder'], f"{stock_code}.html")
                
                if not Config.OVERWRITE_EXISTING and os.path.exists(html_output_file):
                    print(f"⏭️  個別檔案已存在，跳過: {stock_code}")
                else:
                    ChartPlotly.generate_chart(
                        result,
                        stock_code,
                        stock_name,
                        html_output_path=html_output_file
                    )
                    print(f"✅ 個別圖表: {os.path.basename(config['html_output_folder'])}/{stock_code}.html")
            
            print(f"✅ 圖表已生成")
            return html_string

        except Exception as e:
            print(f"❌ 圖表生成失敗: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    def batch_process_all_stocks(base_dir, config):
        """批次處理所有股票 - 按照買超排名順序生成"""

        print("\n" + "="*70)
        print(f"批次處理模式 - {config['market_name']}")
        print(f"輸出方式: 個別HTML + 合併HTML (按買超排名排序)")
        print(f"覆蓋模式: {'覆蓋已存在檔案' if Config.OVERWRITE_EXISTING else '跳過已存在檔案'}")
        print("="*70)

        # 讀取買超排名順序
        ranking_file = os.path.join(config['merged_output_folder'], f"{config['market_type']}_buy_ranking.txt")
        ranked_stocks = []
        
        if os.path.exists(ranking_file):
            print(f"\n✓ 找到買超排名檔案: {os.path.basename(ranking_file)}")
            with open(ranking_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                for line in lines[1:]:  # 跳過標題行
                    parts = line.strip().split(',')
                    if len(parts) >= 2:
                        ranked_stocks.append(parts[1])  # 證券代號
            print(f"  已載入 {len(ranked_stocks)} 支排名股票")
            if len(ranked_stocks) > 0:
                print(f"  前10名: {', '.join(ranked_stocks[:10])}")
        else:
            print(f"\n⚠ 找不到買超排名檔案: {ranking_file}")
            print("  將使用預設順序處理")

        # 取得所有股票代碼
        print("\n⏳ 掃描歷史資料夾...")
        all_stocks = Utils.get_all_stock_codes_from_history(config['history_folder'])

        if not all_stocks:
            print("❌ 無法取得股票清單")
            return

        # 將股票分為兩組：有排名的 + 其他的
        if ranked_stocks:
            # 確保排名中的股票都存在於歷史資料中
            ranked_stocks = [s for s in ranked_stocks if s in all_stocks]
            # 其他未排名的股票（按代碼排序）
            other_stocks = sorted([s for s in all_stocks if s not in ranked_stocks])
            # 合併：先排名的，後其他的
            stock_codes = ranked_stocks + other_stocks
            print(f"\n✓ 股票處理順序:")
            print(f"  - 買超排名股票: {len(ranked_stocks)} 支 (優先處理)")
            print(f"  - 其他股票: {len(other_stocks)} 支")
        else:
            stock_codes = sorted(all_stocks)
            print(f"\n✓ 找到 {len(stock_codes)} 支股票 (按代碼排序)")

        start_time = datetime.now()

        # 同時生成個別HTML和收集合併HTML
        merged_html_parts = []
        success_count = 0
        fail_count = 0

        for idx, stock_code in enumerate(stock_codes, 1):
            print(f"\n{'='*70}")
            
            # 顯示是否為排名股票
            if ranked_stocks and stock_code in ranked_stocks:
                rank = ranked_stocks.index(stock_code) + 1
                print(f"進度: [{idx}/{len(stock_codes)}] 📊 買超排名 #{rank}")
            else:
                print(f"進度: [{idx}/{len(stock_codes)}] ({idx/len(stock_codes)*100:.1f}%)")
            
            print(f"{'='*70}")

            html_string = Processor.process_stock(
                stock_code,
                base_dir,  # 修改：改用 base_dir
                config,
                save_individual=True  # 同時儲存個別檔案
            )

            if html_string:
                merged_html_parts.append(html_string)
                # 在每個圖表之間加入分隔線
                if idx < len(stock_codes):
                    merged_html_parts.append('<div class="stock-separator"></div>')
                success_count += 1
            else:
                fail_count += 1

        # 生成合併的 HTML
        if merged_html_parts:
            print(f"\n{'='*70}")
            print("⏳ 生成合併HTML...")
            print(f"{'='*70}")

            all_charts_html = '\n'.join(merged_html_parts)

            # 包裝成完整的 HTML
            full_html = ChartPlotly._wrap_html(
                all_charts_html,
                f"{config['market_name']}股票技術分析圖表合集"
            )

            # 儲存合併後的 HTML 到 StockInfo 資料夾
            merged_filename = f"ALL_{config['market_type']}.html"
            merged_output_path = os.path.join(config['merged_output_folder'], merged_filename)

            with open(merged_output_path, 'w', encoding='utf-8') as f:
                f.write(full_html)

            print(f"\n✅ 合併HTML已儲存!")
            print(f"  檔案: {merged_filename}")
            print(f"  路徑: {merged_output_path}")
            print(f"  檔案大小: {os.path.getsize(merged_output_path) / 1024 / 1024:.2f} MB")

        end_time = datetime.now()
        elapsed_time = (end_time - start_time).total_seconds()

        print("\n" + "="*70)
        print("批次處理完成")
        print("="*70)
        print(f"總股票數: {len(stock_codes)}")
        print(f"成功處理: {success_count}")
        print(f"處理失敗: {fail_count}")
        print(f"處理時間: {elapsed_time:.1f} 秒 ({elapsed_time/60:.1f} 分鐘)")
        print("="*70)
        print(f"個別HTML位置: {config['html_output_folder']}")
        print(f"合併HTML位置: {config['merged_output_folder']}")
        print("="*70)

def run_step3_chart_generation(base_dir, market_type):
    """執行第三步：圖表生成"""
    print(f"\n{'🔥'*40}")
    print(f"第三步圖表生成：{market_type} ({'上市' if market_type == 'TSE' else '上櫃'})")
    print(f"{'🔥'*40}\n")
    
    # 設定配置
    config = Config.setup_config(base_path=base_dir, market_type=market_type)
    
    # 根據 TOP_STOCKS_ONLY 決定資料夾路徑
    if not TOP_STOCKS_ONLY:
        # 使用 local_ 開頭的資料夾
        if market_type == 'TSE':
            config['history_folder'] = os.path.join(base_dir, 'local_StockTSEHistory')
            config['html_output_folder'] = os.path.join(base_dir, 'local_StockTSEHTML')
        else:
            config['history_folder'] = os.path.join(base_dir, 'local_StockOTCHistory')
            config['html_output_folder'] = os.path.join(base_dir, 'local_StockOTCHTML')
        
        # 確保資料夾存在
        os.makedirs(config['history_folder'], exist_ok=True)
        os.makedirs(config['html_output_folder'], exist_ok=True)
        print(f"📁 History 資料夾: {config['history_folder']}")
        print(f"📁 HTML 資料夾: {config['html_output_folder']}\n")
    # 設定字體 (GitHub Actions 環境)
    Utils.setup_chinese_font(base_dir)
    
    # 批次處理所有股票
    Processor.batch_process_all_stocks(base_dir, config)
    
    print(f"\n✓ {market_type} 圖表生成完成")

# ============================================================================
# 主程式流程
# ============================================================================

def copy_data_to_repo(base_dir, repo_data_dir='data'):
    """
    將下載和處理的資料複製到 repo 的 data 資料夾
    
    Args:
        base_dir: 工作目錄
        repo_data_dir: repo 中的 data 資料夾路徑
    """
    print("\n" + "📦"*40)
    print("複製資料到 Repository")
    print("📦"*40 + "\n")
    
    # 確保 repo data 目錄存在
    os.makedirs(repo_data_dir, exist_ok=True)
    
    # 定義需要複製的資料夾
    folders_to_copy = [
        'StockTSEDaily',      # 上市每日交易
        'StockTSEShares',     # 上市三大法人
        'StockOTCDaily',   # 上櫃每日交易
        'StockOTCShares',  # 上櫃三大法人
        'StockTSEHistory',    # 上市歷史資料
        'StockOTCHistory', # 上櫃歷史資料
        'StockInfo',       # 分析報告
        'StockTSEHTML',       # 上市圖表 HTML
        'StockOTCHTML'    # 上櫃圖表 HTML
    ]
    
    copied_count = 0
    skipped_count = 0
    
    for folder_name in folders_to_copy:
        source_path = os.path.join(base_dir, folder_name)
        dest_path = os.path.join(repo_data_dir, folder_name)
        
        if os.path.exists(source_path):
            try:
                # 如果目標資料夾存在,先刪除
                if os.path.exists(dest_path):
                    shutil.rmtree(dest_path)
                
                # 複製整個資料夾
                shutil.copytree(source_path, dest_path)
                
                # 計算檔案數量
                file_count = len([f for f in os.listdir(dest_path) if os.path.isfile(os.path.join(dest_path, f))])
                print(f"✓ {folder_name:<20} → {file_count} 個檔案")
                copied_count += 1
            except Exception as e:
                print(f"✗ {folder_name:<20} 複製失敗: {e}")
        else:
            print(f"⊘ {folder_name:<20} 來源不存在")
            skipped_count += 1
    
    print("\n" + "="*80)
    print(f"複製完成: {copied_count} 個資料夾, 跳過: {skipped_count} 個")
    print("="*80 + "\n")

def main():
    """主程式 - 完整自動化流程"""
    
    # 解析命令列參數
    parser = argparse.ArgumentParser(description='台灣股市資料完整處理流程')
    parser.add_argument('--base-dir', type=str, default=None,
                       help='指定工作目錄 (預設: 當前目錄)')
    parser.add_argument('--repo-data-dir', type=str, default='data',
                       help='Repository 的 data 資料夾路徑 (預設: data)')
    parser.add_argument('--copy-to-repo', action='store_true',
                       help='完成後將資料複製到 repo 的 data 資料夾')
    parser.add_argument('--start-date', type=str, default='2025-01-01',
                       help='爬蟲起始日期 (格式: YYYY-MM-DD)')
    parser.add_argument('--skip-crawler', action='store_true',
                       help='跳過爬蟲步驟')
    parser.add_argument('--skip-analysis', action='store_true',
                       help='跳過分析步驟')
    parser.add_argument('--skip-charts', action='store_true',
                       help='跳過圖表生成步驟')
    parser.add_argument('--market', type=str, choices=['TSE', 'OTC', 'BOTH'], 
                       default='BOTH', help='處理市場類型')
    parser.add_argument('--debug-skip-data-processing', action='store_true',
                       help='除錯模式：跳過爬蟲和 History 生成,直接測試報表和上傳')
    
    args = parser.parse_args()
    
    print("\n" + "="*80)
    print("台灣股市資料完整處理流程 - GitHub Actions 版本")
    print("="*80)
    print("流程說明：")
    if not args.skip_crawler:
        print("  1. 執行爬蟲程式 (上市/上櫃每日交易與三大法人)")
    if not args.skip_analysis:
        print("  2. 清理舊的 History 資料夾")
        print("  3. 執行分析程式 - TSE (上市)" if args.market in ['TSE', 'BOTH'] else "")
        print("  4. 執行分析程式 - OTC (上櫃)" if args.market in ['OTC', 'BOTH'] else "")
    if not args.skip_charts:
        print("  5. 清理舊的圖表資料夾")
        print("  6. 執行圖表生成 - TSE (上市)" if args.market in ['TSE', 'BOTH'] else "")
        print("  7. 執行圖表生成 - OTC (上櫃)" if args.market in ['OTC', 'BOTH'] else "")
    print("="*80 + "\n")
    
    # 設定基礎目錄
    if args.base_dir:
        os.environ['STOCK_DATA_DIR'] = args.base_dir
    base_dir = setup_base_directory()
    
    # 建立必要的資料夾結構
    create_required_directories(base_dir)
    
    # ========== 步驟 1：爬蟲 ==========
    if args.debug_skip_data_processing:
        print("\n" + "⚡"*40)
        print("除錯模式：跳過爬蟲和 History 生成步驟")
        print("⚡"*40 + "\n")
    else:
        if not args.skip_crawler:
            start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
            run_step1_crawler(base_dir, start_date=start_date)
        
        # ========== 步驟 2-4：分析 ==========
        if not args.skip_analysis:
            # 刪除 History 資料夾
            print("\n" + "🔥"*40)
            print("步驟 2：清理 History 資料夾")
            print("🔥"*40)
            # 根據 TOP_STOCKS_ONLY 決定要清理的資料夾
            if TOP_STOCKS_ONLY:
                delete_folders(base_dir, ['StockTSEHistory', 'StockOTCHistory'])
            else:
                delete_folders(base_dir, ['local_StockTSEHistory', 'local_StockOTCHistory'])
            
            # 執行分析
            if args.market in ['TSE', 'BOTH']:
                run_step2_analysis(base_dir, 'TSE')
            
            if args.market in ['OTC', 'BOTH']:
                run_step2_analysis(base_dir, 'OTC')
    
    # ========== 步驟 5-7：圖表生成 ==========
    if not args.skip_charts:
        # 刪除圖表資料夾
        print("\n" + "🔥"*40)
        print("步驟 5：清理圖表資料夾")
        print("🔥"*40)
        # 根據 TOP_STOCKS_ONLY 決定要清理的資料夾
        if TOP_STOCKS_ONLY:
            delete_folders(base_dir, ['StockTSEHTML', 'StockOTCHTML'])
        else:
            delete_folders(base_dir, ['local_StockTSEHTML', 'local_StockOTCHTML'])
        
        # 執行圖表生成
        if args.market in ['TSE', 'BOTH']:
            run_step3_chart_generation(base_dir, 'TSE')
        
        if args.market in ['OTC', 'BOTH']:
            run_step3_chart_generation(base_dir, 'OTC')
    
    # ========== 步驟 7.5：複製帶日期的檔案 ==========
    print("\n" + "📅"*40)
    print("步驟 7.5：備份帶日期的分析檔案")
    print("📅"*40 + "\n")
    
    
    stock_info_dir = os.path.join(base_dir, 'StockInfo')
    
    # 先從 Excel 檔案提取日期
    def extract_date_from_excel(excel_file_path):
        """從 Excel 檔案的第二個分頁名稱提取日期"""
        try:
            wb = load_workbook(excel_file_path)
            sheet_names = wb.sheetnames
            if len(sheet_names) >= 2:
                second_sheet_name = sheet_names[1]
                match = re.search(r'(\d{8})', second_sheet_name)
                wb.close()
                if match:
                    return match.group(1)
            wb.close()
        except Exception as e:
            print(f"  ⚠️  提取日期失敗: {e}")
        return None
    
    # 先備份 Excel 檔案並提取日期
    tse_date_str = None
    otc_date_str = None
    
    excel_files_to_backup = [
        ('tse_analysis_result.xlsx', 'TSE'),
        ('otc_analysis_result.xlsx', 'OTC'),
    ]
    
    backup_count = 0
    
    for source_name, market_type in excel_files_to_backup:
        source_path = os.path.join(stock_info_dir, source_name)
        
        if os.path.exists(source_path):
            # 提取日期
            date_str = extract_date_from_excel(source_path)
            
            if date_str:
                # 儲存日期供後續 HTML 使用
                if market_type == 'TSE':
                    tse_date_str = date_str
                else:
                    otc_date_str = date_str
                
                # 備份 Excel
                backup_name = f'{source_name.replace(".xlsx", "")}_{date_str}.xlsx'
                backup_path = os.path.join(stock_info_dir, backup_name)
                
                try:
                    shutil.copy2(source_path, backup_path)
                    file_size = os.path.getsize(backup_path) / 1024  # KB
                    print(f"✅ 已備份: {source_name} → {backup_name} ({file_size:.1f} KB, 日期: {date_str})")
                    backup_count += 1
                except Exception as e:
                    print(f"❌ 備份失敗: {source_name} - {e}")
            else:
                print(f"⚠️  無法從 {source_name} 提取日期,使用當前日期")
                # 如果無法提取日期,使用台灣時間
                from datetime import timezone, timedelta as td
                taiwan_tz = timezone(td(hours=8))
                taiwan_time = datetime.now(taiwan_tz)
                date_str = taiwan_time.strftime('%Y%m%d')
                
                if market_type == 'TSE':
                    tse_date_str = date_str
                else:
                    otc_date_str = date_str
        else:
            print(f"⚠️  Excel 檔案不存在: {source_name}")
    
    # 使用提取的日期備份 HTML
    html_files_to_backup = [
        ('ALL_TSE.html', tse_date_str),
        ('ALL_OTC.html', otc_date_str),
    ]
    
    for source_name, date_str in html_files_to_backup:
        if date_str:
            source_path = os.path.join(stock_info_dir, source_name)
            backup_name = f'{source_name.replace(".html", "")}_{date_str}.html'
            backup_path = os.path.join(stock_info_dir, backup_name)
            
            if os.path.exists(source_path):
                try:
                    shutil.copy2(source_path, backup_path)
                    file_size = os.path.getsize(backup_path) / 1024  # KB
                    print(f"✅ 已備份: {source_name} → {backup_name} ({file_size:.1f} KB, 日期: {date_str})")
                    backup_count += 1
                except Exception as e:
                    print(f"❌ 備份失敗: {source_name} - {e}")
            else:
                print(f"⚠️  HTML 檔案不存在: {source_name}")
        else:
            print(f"⚠️  無法取得 {source_name} 的日期,跳過備份")
    
    print(f"\n✓ 共備份 {backup_count} 個檔案")
    print("="*80 + "\n")
    
    # ========== 步驟 7.6：清理 Excel 分頁 ==========
    print("\n" + "📝"*40)
    print("步驟 7.6：清理 Excel 分頁（只保留最近交易日）")
    print("📝"*40 + "\n")
    
    # 處理帶日期的 Excel 檔案 - 使用從 Excel 提取的日期
    excel_files_to_clean = []
    if tse_date_str:
        excel_files_to_clean.append(f'tse_analysis_result_{tse_date_str}.xlsx')
    if otc_date_str:
        excel_files_to_clean.append(f'otc_analysis_result_{otc_date_str}.xlsx')

    
    cleaned_count = 0
    for excel_file in excel_files_to_clean:
        excel_path = os.path.join(stock_info_dir, excel_file)
        
        if os.path.exists(excel_path):
            print(f"處理檔案: {excel_file}")
            result = clean_excel_keep_second_sheet(excel_path)
            if result:
                cleaned_count += 1
                print()
        else:
            print(f"⊘ 檔案不存在: {excel_file}\n")
    
    print(f"✓ 共處理 {cleaned_count} 個 Excel 檔案")
    print("="*80 + "\n")
    
    # ========== 步驟 8：複製到 Repository ==========
    if args.copy_to_repo:
        copy_data_to_repo(base_dir, args.repo_data_dir)
    
    # ========== 完成 ==========
    print("\n" + "🎉"*40)
    print("所有流程已完成！")
    print("🎉"*40 + "\n")
    
    print("處理結果：")
    if not args.skip_crawler:
        print("  ✓ 上市/上櫃每日交易資料已更新")
        print("  ✓ 三大法人買賣超資料已更新")
    if not args.skip_analysis:
        if args.market in ['TSE', 'BOTH']:
            print("  ✓ TSE 分析報告 (Excel) 已生成")
        if args.market in ['OTC', 'BOTH']:
            print("  ✓ OTC 分析報告 (Excel) 已生成")
    if not args.skip_charts:
        if args.market in ['TSE', 'BOTH']:
            print("  ✓ TSE 技術分析圖表 (HTML) 已生成")
        if args.market in ['OTC', 'BOTH']:
            print("  ✓ OTC 技術分析圖表 (HTML) 已生成")
    print("\n" + "="*80)

if __name__ == "__main__":
    main()