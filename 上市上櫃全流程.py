#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台灣股市資料完整處理流程 - GitHub Actions 版
"""

import os
import glob
import shutil
import requests
import pandas as pd
import numpy as np
import time
from datetime import datetime, timedelta
from io import StringIO
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# 基礎目錄
BASE_DIR = os.getcwd()

def delete_folders(folder_names):
    """刪除指定的資料夾"""
    print(f"\n{'='*80}")
    print("清理資料夾...")
    print(f"{'='*80}")
    for folder_name in folder_names:
        folder_path = os.path.join(BASE_DIR, folder_name)
        if os.path.exists(folder_path):
            try:
                shutil.rmtree(folder_path)
                print(f"✓ 已刪除: {folder_name}")
            except Exception as e:
                print(f"✗ 刪除失敗 {folder_name}: {e}")
        else:
            print(f"⊘ 資料夾不存在: {folder_name}")
    print(f"{'='*80}\n")


# ============================================================================
# 第一步:爬蟲程式
# ============================================================================

def filter_csv_content(csv_bytes):
    """過濾 CSV 內容,只保留股票資料"""
    try:
        content = csv_bytes.decode('cp950')
        lines = content.split('\r\n')

        filtered_lines = []
        header_found = False
        stock_count = 0

        for line in lines:
            if '證券代號' in line and not header_found:
                filtered_lines.append(line)
                header_found = True
                continue

            if header_found:
                match = re.match(r'^=?"?(\d{4})"?', line)
                if match:
                    filtered_lines.append(line)
                    stock_count += 1

        filtered_content = '\r\n'.join(filtered_lines)
        filtered_bytes = filtered_content.encode('cp950')
        print(f"   ✂️  過濾完成:保留 {stock_count} 檔股票")
        return filtered_bytes

    except Exception as e:
        print(f"   ⚠️  過濾失敗: {e},將儲存原始資料")
        return csv_bytes

def download_twse_daily(date_str):
    """下載上市每日交易資料"""
    if '-' in date_str:
        date_str = date_str.replace('-', '')

    url = f"https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date={date_str}&type=ALL&response=csv"

    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200 and len(response.content) > 100:
            return response.content
        return None
    except Exception as e:
        print(f"   ❌ 下載錯誤: {e}")
        return None

def crawl_twse_daily(start_date, end_date, save_dir):
    """抓取上市每日交易資料"""
    print("="*60)
    print("📊 [1/4] 上市每日交易資料 (TWSE Daily)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    # 從今天往回檢查
    while curr >= start_date:
        if curr.weekday() < 5:  # 只檢查平日
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                print(f"  {date_formatted}... [已存在,停止檢查] ✓")
                break
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        csv_bytes = download_twse_daily(date_str)

        if csv_bytes:
            filtered_bytes = filter_csv_content(csv_bytes)
            with open(file_path, 'wb') as f:
                f.write(filtered_bytes)
            print(" ✓")
            success_count += 1
        else:
            print(" ✗")

        time.sleep(1)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count

# ============================================================================
# 2. 上市三大法人買賣超 (TWSE Institutional)
# ============================================================================

def download_twse_institutional(date_str):
    """下載上市三大法人資料"""
    url = 'https://www.twse.com.tw/rwd/zh/fund/T86'
    params = {'date': date_str, 'selectType': 'ALL', 'response': 'json'}
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    try:
        response = requests.get(url, params=params, headers=headers, timeout=30)
        response.raise_for_status()
        data = response.json()

        if data.get('stat') == 'OK' and 'data' in data:
            return pd.DataFrame(data['data'], columns=data['fields'])
        return None
    except Exception as e:
        print(f"   ❌ 錯誤: {e}")
        return None

def crawl_twse_institutional(start_date, end_date, save_dir):
    """抓取上市三大法人買賣超資料"""
    print("="*60)
    print("📊 [2/4] 上市三大法人買賣超 (TWSE Institutional)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    while curr >= start_date:
        if curr.weekday() < 5:
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                print(f"  {date_formatted}... [已存在,停止檢查] ✓")
                break
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        df = download_twse_institutional(date_str)

        if df is not None and not df.empty:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(" ✓")
            success_count += 1
        else:
            print(" ✗")

        time.sleep(3)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count

# ============================================================================
# 3. 上櫃每日交易資料 (OTC Daily)
# ============================================================================

def process_otc_daily_columns(df):
    """處理上櫃每日交易資料欄位"""
    rename_mapping = {
        '代號': '證券代號',
        '名稱': '證券名稱',
        '收盤': '收盤價',
        '開盤': '開盤價',
        '最高': '最高價',
        '最低': '最低價',
        '成交股數': '成交股數',
        '成交筆數': '成交筆數',
        '成交金額(元)': '成交金額',
        '漲跌': '漲跌價差',
        '最後買價': '最後揭示買價',
        '最後買量(千股)': '最後揭示買量',
        '最後賣價': '最後揭示賣價',
        '最後賣量(千股)': '最後揭示賣量'
    }

    df = df.rename(columns=rename_mapping)

    numeric_columns = ['收盤價', '開盤價', '最高價', '最低價',
                      '成交股數', '成交筆數', '成交金額']

    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')

    if '漲跌價差' in df.columns:
        def parse_change(val):
            if pd.isna(val) or val == '':
                return 0
            val_str = str(val).replace(',', '').strip()
            if val_str == '-' or val_str == '除權息' or val_str == '除息' or val_str == '除權':
                return 0
            try:
                return float(val_str)
            except:
                return 0

        df['漲跌價差'] = df['漲跌價差'].apply(parse_change)

    return df

def download_otc_daily(date_str, max_retries=3):
    """下載上櫃每日交易資料"""
    if '-' in date_str:
        date_str = date_str.replace('-', '')

    minguo_year = str(int(date_str[:4]) - 1911)
    date_formatted = f"{minguo_year}/{date_str[4:6]}/{date_str[6:8]}"
    url = f"https://www.tpex.org.tw/web/stock/aftertrading/otc_quotes_no1430/stk_wn1430_result.php?l=zh-tw&d={date_formatted}"

    for attempt in range(max_retries):
        try:
            response = requests.get(url, timeout=30)
            if response.status_code == 200:
                data = response.json()
                if 'aaData' in data and len(data['aaData']) > 0:
                    df = pd.DataFrame(data['aaData'])
                    if len(df.columns) >= 14:
                        df.columns = ['代號', '名稱', '收盤', '漲跌', '開盤', '最高', '最低',
                                    '成交股數', '成交金額(元)', '成交筆數', '最後買價',
                                    '最後買量(千股)', '最後賣價', '最後賣量(千股)']
                        return process_otc_daily_columns(df)
            return None
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2)
                continue
            print(f"   ❌ 錯誤: {e}")
            return None

def crawl_otc_daily(start_date, end_date, save_dir):
    """抓取上櫃每日交易資料"""
    print("="*60)
    print("📊 [3/4] 上櫃每日交易資料 (OTC Daily)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    while curr >= start_date:
        if curr.weekday() < 5:
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                print(f"  {date_formatted}... [已存在,停止檢查] ✓")
                break
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        df = download_otc_daily(date_str)

        if df is not None and not df.empty:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(" ✓")
            success_count += 1
        else:
            print(" ✗")

        time.sleep(3)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count

# ============================================================================
# 4. 上櫃三大法人買賣超 (OTC Institutional)
# ============================================================================

def download_otc_institutional(date_str):
    """下載上櫃三大法人資料"""
    if '-' in date_str:
        date_str = date_str.replace('-', '')

    minguo_year = str(int(date_str[:4]) - 1911)
    date_formatted = f"{minguo_year}/{date_str[4:6]}/{date_str[6:8]}"
    url = "https://www.tpex.org.tw/web/stock/3insti/daily_trade/3itrade_hedge_result.php"
    params = {'l': 'zh-tw', 'd': date_formatted, 't': 'D'}
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    try:
        response = requests.get(url, params=params, headers=headers, timeout=30)
        response.raise_for_status()
        data = response.json()

        if 'aaData' in data and len(data['aaData']) > 0:
            df = pd.DataFrame(data['aaData'])
            columns = ['代號', '名稱', '外資及陸資(不含外資自營商)-買進股數', '外資及陸資(不含外資自營商)-賣出股數',
                      '外資及陸資(不含外資自營商)-買賣超股數', '外資自營商-買進股數', '外資自營商-賣出股數',
                      '外資自營商-買賣超股數', '投信-買進股數', '投信-賣出股數', '投信-買賣超股數',
                      '自營商-買進股數(自行買賣)', '自營商-賣出股數(自行買賣)', '自營商-買賣超股數(自行買賣)',
                      '自營商-買進股數(避險)', '自營商-賣出股數(避險)', '自營商-買賣超股數(避險)',
                      '三大法人買賣超股數']

            if len(df.columns) == len(columns):
                df.columns = columns
                return df

        return None
    except Exception as e:
        print(f"   ❌ 錯誤: {e}")
        return None

def crawl_otc_institutional(start_date, end_date, save_dir):
    """抓取上櫃三大法人買賣超資料"""
    print("="*60)
    print("📊 [4/4] 上櫃三大法人買賣超 (OTC Institutional)")
    print("="*60)

    os.makedirs(save_dir, exist_ok=True)

    missing_dates = []
    curr = end_date

    while curr >= start_date:
        if curr.weekday() < 5:
            date_formatted = curr.strftime('%Y-%m-%d')
            file_path = os.path.join(save_dir, f'{date_formatted}.csv')

            if os.path.exists(file_path):
                print(f"  {date_formatted}... [已存在,停止檢查] ✓")
                break
            else:
                missing_dates.append(curr)

        curr -= timedelta(days=1)

    if not missing_dates:
        print("✓ 無缺失資料\n")
        return 0

    print(f"需要下載 {len(missing_dates)} 個交易日")
    print("-"*60)

    success_count = 0

    for idx, date_dt in enumerate(missing_dates, 1):
        date_str = date_dt.strftime('%Y%m%d')
        date_formatted = date_dt.strftime('%Y-%m-%d')
        file_path = os.path.join(save_dir, f'{date_formatted}.csv')

        print(f"  [{idx:2d}/{len(missing_dates)}] {date_formatted}...", end='', flush=True)

        df = download_otc_institutional(date_str)

        if df is not None and not df.empty:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(" ✓")
            success_count += 1
        else:
            print(" ✗")

        time.sleep(3)

    print(f"✓ 成功下載: {success_count} 個檔案\n")
    return success_count


# ============================================================================
# 第二步:分析程式 - 輔助函數
# ============================================================================

def setup_config(market_type='TSE'):
    """設定配置"""
    if market_type == 'TSE':
        return {
            'market_type': market_type,
            'market_name': '上市',
            'folder_path': os.path.join(BASE_DIR, 'StockShares'),
            'stock_daily_folder': os.path.join(BASE_DIR, 'StockDaily'),
            'history_folder': os.path.join(BASE_DIR, 'StockHistory'),
            'market_list_path': os.path.join(BASE_DIR, 'StockList', 'stockListTSE.csv'),
            'output_path': os.path.join(BASE_DIR, 'TSE.xlsx'),
            'sigma_threshold': 2,
            'aggregate_threshold': 10000,
            'show_top_n': None
        }
    else:  # OTC
        return {
            'market_type': market_type,
            'market_name': '上櫃',
            'folder_path': os.path.join(BASE_DIR, 'StockOTCShares'),
            'stock_daily_folder': os.path.join(BASE_DIR, 'StockOTCDaily'),
            'history_folder': os.path.join(BASE_DIR, 'StockOTCHistory'),
            'market_list_path': os.path.join(BASE_DIR, 'StockList', 'stockListOTC.csv'),
            'output_path': os.path.join(BASE_DIR, 'OTC.xlsx'),
            'sigma_threshold': 2,
            'aggregate_threshold': 10000,
            'show_top_n': None
        }

def load_stock_list(filepath):
    """載入股票清單"""
    try:
        df = pd.read_csv(filepath, encoding='utf-8-sig')
        allowed_stock_codes = set(df['代號'].astype(str))

        stock_sector_map = dict(zip(df['代號'].astype(str), df['產業別']))

        etf_stock_codes = set()
        if '產業別' in df.columns:
            etf_df = df[df['產業別'].str.contains('ETF', na=False)]
            etf_stock_codes = set(etf_df['代號'].astype(str))

        return allowed_stock_codes, stock_sector_map, etf_stock_codes
    except Exception as e:
        print(f"❌ 無法載入股票清單: {e}")
        return set(), {}, set()

def load_stock_daily_prices(folder_path, allowed_stock_codes):
    """載入股票每日價格"""
    stock_daily_prices = {}

    if not os.path.exists(folder_path):
        print(f"⚠️  每日價格資料夾不存在: {folder_path}")
        return stock_daily_prices

    csv_files = glob.glob(os.path.join(folder_path, '*.csv'))

    if not csv_files:
        print(f"⚠️  找不到每日價格資料")
        return stock_daily_prices

    latest_file = max(csv_files, key=os.path.getmtime)

    try:
        df = pd.read_csv(latest_file, encoding='cp950', dtype=str)

        if '證券代號' in df.columns and '收盤價' in df.columns:
            df = df[df['證券代號'].isin(allowed_stock_codes)]
            df['收盤價'] = df['收盤價'].str.replace(',', '').str.replace('+', '').str.replace('-', '')
            df['收盤價'] = pd.to_numeric(df['收盤價'], errors='coerce')
            stock_daily_prices = dict(zip(df['證券代號'], df['收盤價']))

    except Exception as e:
        print(f"⚠️  讀取每日價格失敗: {e}")

    return stock_daily_prices

def get_latest_files(folder_path, num_files=61):
    """取得最近的檔案"""
    csv_files = glob.glob(os.path.join(folder_path, '*.csv'))

    if not csv_files:
        return []

    csv_files_with_time = [(f, os.path.getmtime(f)) for f in csv_files]
    csv_files_with_time.sort(key=lambda x: x[1], reverse=True)
    latest_files = [f for f, _ in csv_files_with_time[:num_files]]

    return latest_files

def read_shares_file(filepath):
    """讀取單一三大法人檔案"""
    try:
        df = pd.read_csv(filepath, encoding='utf-8-sig')
        return df
    except Exception as e:
        print(f"❌ 讀取失敗 {os.path.basename(filepath)}: {e}")
        return pd.DataFrame()

def process_shares_files(latest_files, allowed_stock_codes, stock_daily_prices,
                        stock_sector_map, etf_stock_codes):
    """處理三大法人檔案"""
    all_data = []
    daily_buy_sell_data = {}
    etf_daily_data = {}
    buy_top20_tracker = {}
    sell_top20_tracker = {}
    daily_buy_stocks = {}
    daily_sell_stocks = {}
    daily_all_stocks = {}
    all_historical_data = {}
    statistics = {'processed': 0, 'skipped': 0}

    for filepath in latest_files:
        date_str = os.path.basename(filepath).replace('.csv', '')

        df = read_shares_file(filepath)

        if df.empty:
            statistics['skipped'] += 1
            continue

        if '證券代號' not in df.columns or '三大法人買賣超股數' not in df.columns:
            statistics['skipped'] += 1
            continue

        df = df[df['證券代號'].isin(allowed_stock_codes)].copy()

        df['三大法人買賣超股數'] = df['三大法人買賣超股數'].astype(str).str.replace(',', '')
        df['三大法人買賣超股數'] = pd.to_numeric(df['三大法人買賣超股數'], errors='coerce')
        df = df.dropna(subset=['三大法人買賣超股數'])

        df['產業別'] = df['證券代號'].map(stock_sector_map)
        df['收盤價'] = df['證券代號'].map(stock_daily_prices)

        # 記錄所有股票
        daily_all_stocks[date_str] = set(df['證券代號'])

        # 買超
        buy_df = df[df['三大法人買賣超股數'] > 0].copy()
        daily_buy_stocks[date_str] = set(buy_df['證券代號'])

        # ETF 買超
        etf_buy_df = buy_df[buy_df['證券代號'].isin(etf_stock_codes)].copy()
        if not etf_buy_df.empty:
            etf_buy_df = etf_buy_df.sort_values('三大法人買賣超股數', ascending=False).head(50)
            etf_daily_data[date_str] = {'buy': etf_buy_df}

        # 非 ETF 買超
        buy_df = buy_df[~buy_df['證券代號'].isin(etf_stock_codes)]

        if not buy_df.empty:
            buy_df = buy_df.sort_values('三大法人買賣超股數', ascending=False).head(50)
            daily_buy_sell_data[date_str] = {'buy': buy_df}

            top20_buy = set(buy_df.head(20)['證券代號'])
            for stock_code in top20_buy:
                if stock_code not in buy_top20_tracker:
                    buy_top20_tracker[stock_code] = []
                buy_top20_tracker[stock_code].append(date_str)

        # 賣超
        sell_df = df[df['三大法人買賣超股數'] < 0].copy()
        daily_sell_stocks[date_str] = set(sell_df['證券代號'])

        sell_df = sell_df[~sell_df['證券代號'].isin(etf_stock_codes)]

        if not sell_df.empty:
            sell_df = sell_df.sort_values('三大法人買賣超股數').head(50)

            if date_str in daily_buy_sell_data:
                daily_buy_sell_data[date_str]['sell'] = sell_df
            else:
                daily_buy_sell_data[date_str] = {'sell': sell_df}

            top20_sell = set(sell_df.head(20)['證券代號'])
            for stock_code in top20_sell:
                if stock_code not in sell_top20_tracker:
                    sell_top20_tracker[stock_code] = []
                sell_top20_tracker[stock_code].append(date_str)

        all_data.append(df)

        # 歷史資料
        for _, row in df.iterrows():
            stock_code = row['證券代號']
            if stock_code not in all_historical_data:
                all_historical_data[stock_code] = []
            all_historical_data[stock_code].append(row['三大法人買賣超股數'])

        statistics['processed'] += 1

    return (all_data, daily_buy_sell_data, etf_daily_data, buy_top20_tracker,
            sell_top20_tracker, daily_buy_stocks, daily_sell_stocks,
            daily_all_stocks, all_historical_data, statistics)

def calculate_stock_statistics(all_historical_data, sigma_threshold):
    """計算股票統計數據"""
    stock_statistics = {}

    for stock_code, values in all_historical_data.items():
        mean = np.mean(values)
        std = np.std(values, ddof=1)

        stock_statistics[stock_code] = {
            'mean': mean,
            'std': std,
            'upper_threshold': mean + sigma_threshold * std,
            'lower_threshold': mean - sigma_threshold * std
        }

    return stock_statistics

def analyze_new_entries_and_observables(latest_file, daily_buy_stocks, daily_sell_stocks,
                                       daily_all_stocks, stock_statistics,
                                       allowed_stock_codes, sigma_threshold):
    """分析新進榜與觀察股"""
    latest_date = os.path.basename(latest_file).replace('.csv', '')
    sorted_dates = sorted(daily_buy_stocks.keys(), reverse=True)

    if len(sorted_dates) < 2:
        return set(), set(), set(), set(), latest_date, [], []

    latest_buy = daily_buy_stocks.get(sorted_dates[0], set())
    latest_sell = daily_sell_stocks.get(sorted_dates[0], set())

    previous_dates = sorted_dates[1:6]
    previous_buy = set()
    previous_sell = set()

    for date in previous_dates:
        previous_buy.update(daily_buy_stocks.get(date, set()))
        previous_sell.update(daily_sell_stocks.get(date, set()))

    new_buy_stocks = latest_buy - previous_buy
    new_sell_stocks = latest_sell - previous_sell

    # 觀察股
    observable_buy_stocks = set()
    observable_sell_stocks = set()

    df = read_shares_file(latest_file)
    if not df.empty and '證券代號' in df.columns and '三大法人買賣超股數' in df.columns:
        df = df[df['證券代號'].isin(allowed_stock_codes)]
        df['三大法人買賣超股數'] = pd.to_numeric(
            df['三大法人買賣超股數'].astype(str).str.replace(',', ''),
            errors='coerce'
        )
        df = df.dropna(subset=['三大法人買賣超股數'])

        for _, row in df.iterrows():
            stock_code = row['證券代號']
            value = row['三大法人買賣超股數']

            if stock_code in stock_statistics:
                stats = stock_statistics[stock_code]
                if value > stats['upper_threshold'] and stock_code not in new_buy_stocks:
                    observable_buy_stocks.add(stock_code)
                elif value < stats['lower_threshold'] and stock_code not in new_sell_stocks:
                    observable_sell_stocks.add(stock_code)

    # 買超前 50
    latest_buy_stocks_50 = []
    if not df.empty:
        buy_df = df[df['三大法人買賣超股數'] > 0].copy()
        buy_df = buy_df.sort_values('三大法人買賣超股數', ascending=False).head(50)
        latest_buy_stocks_50 = buy_df['證券代號'].tolist()

    # 賣超前 50
    latest_sell_stocks_50 = []
    if not df.empty:
        sell_df = df[df['三大法人買賣超股數'] < 0].copy()
        sell_df = sell_df.sort_values('三大法人買賣超股數').head(50)
        latest_sell_stocks_50 = sell_df['證券代號'].tolist()

    return (new_buy_stocks, new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
            latest_date, latest_buy_stocks_50, latest_sell_stocks_50)

def collect_stock_history(stock_list, shares_folder, daily_folder, output_folder, allowed_stock_codes):
    """收集股票歷史資料"""
    os.makedirs(output_folder, exist_ok=True)

    for stock_code in stock_list:
        if stock_code not in allowed_stock_codes:
            continue

        output_file = os.path.join(output_folder, f"{stock_code}.csv")

        if os.path.exists(output_file):
            continue

        history_data = []
        shares_files = glob.glob(os.path.join(shares_folder, '*.csv'))

        for filepath in shares_files:
            date_str = os.path.basename(filepath).replace('.csv', '')

            try:
                df = pd.read_csv(filepath, encoding='utf-8-sig')

                if '證券代號' not in df.columns:
                    continue

                stock_df = df[df['證券代號'] == stock_code]

                if stock_df.empty:
                    continue

                for _, row in stock_df.iterrows():
                    history_data.append({
                        '日期': date_str,
                        '證券代號': stock_code,
                        '證券名稱': row.get('證券名稱', ''),
                        '三大法人買賣超股數': row.get('三大法人買賣超股數', 0)
                    })

            except Exception:
                continue

        # 加入每日價格
        daily_files = glob.glob(os.path.join(daily_folder, '*.csv'))

        for filepath in daily_files:
            date_str = os.path.basename(filepath).replace('.csv', '')

            try:
                df = pd.read_csv(filepath, encoding='cp950', dtype=str)

                if '證券代號' not in df.columns or '收盤價' not in df.columns:
                    continue

                stock_df = df[df['證券代號'] == stock_code]

                if stock_df.empty:
                    continue

                close_price = stock_df['收盤價'].iloc[0]
                close_price = close_price.replace(',', '').replace('+', '').replace('-', '')

                try:
                    close_price = float(close_price)
                except:
                    close_price = None

                # 更新歷史資料
                for item in history_data:
                    if item['日期'] == date_str:
                        item['收盤價'] = close_price
                        break

            except Exception:
                continue

        if history_data:
            history_df = pd.DataFrame(history_data)
            history_df = history_df.sort_values('日期')
            history_df.to_csv(output_file, index=False, encoding='utf-8-sig')

def aggregate_analysis(buy_top20_tracker, sell_top20_tracker, stock_sector_map,
                      aggregate_threshold=10000, show_top_n=None):
    """彙整分析"""
    # 買超
    buy_analysis = []

    for stock_code, dates in buy_top20_tracker.items():
        appearance_count = len(dates)
        sector = stock_sector_map.get(stock_code, '未知')

        buy_analysis.append({
            '證券代號': stock_code,
            '產業別': sector,
            '出現次數': appearance_count
        })

    buy_stocks = pd.DataFrame(buy_analysis)

    if not buy_stocks.empty:
        buy_stocks = buy_stocks.sort_values('出現次數', ascending=False)

        if show_top_n:
            buy_stocks = buy_stocks.head(show_top_n)

    # 賣超
    sell_analysis = []

    for stock_code, dates in sell_top20_tracker.items():
        appearance_count = len(dates)
        sector = stock_sector_map.get(stock_code, '未知')

        sell_analysis.append({
            '證券代號': stock_code,
            '產業別': sector,
            '出現次數': appearance_count
        })

    sell_stocks = pd.DataFrame(sell_analysis)

    if not sell_stocks.empty:
        sell_stocks = sell_stocks.sort_values('出現次數', ascending=False)

        if show_top_n:
            sell_stocks = sell_stocks.head(show_top_n)

    # 同時買賣超
    both_stocks_set = set()
    both_stocks_df = pd.DataFrame()

    if not buy_stocks.empty and not sell_stocks.empty:
        buy_set = set(buy_stocks['證券代號'])
        sell_set = set(sell_stocks['證券代號'])
        both_stocks_set = buy_set & sell_set

        if both_stocks_set:
            both_data = []

            for stock_code in both_stocks_set:
                buy_count = buy_stocks[buy_stocks['證券代號'] == stock_code]['出現次數'].values[0]
                sell_count = sell_stocks[sell_stocks['證券代號'] == stock_code]['出現次數'].values[0]
                sector = stock_sector_map.get(stock_code, '未知')

                both_data.append({
                    '證券代號': stock_code,
                    '產業別': sector,
                    '買超次數': buy_count,
                    '賣超次數': sell_count
                })

            both_stocks_df = pd.DataFrame(both_data)
            both_stocks_df = both_stocks_df.sort_values('買超次數', ascending=False)

    return buy_stocks, sell_stocks, both_stocks_set, both_stocks_df

def export_to_excel(output_path, buy_stocks, sell_stocks, both_stocks_set, both_stocks_df,
                   daily_buy_sell_data, etf_daily_data, latest_date, new_buy_stocks,
                   new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
                   stock_sector_map, etf_stock_codes):
    """匯出到 Excel"""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 工作表 1: 彙整買超
        if not buy_stocks.empty:
            buy_stocks.to_excel(writer, sheet_name='彙整買超', index=False)

        # 工作表 2: 彙整賣超
        if not sell_stocks.empty:
            sell_stocks.to_excel(writer, sheet_name='彙整賣超', index=False)

        # 工作表 3: 同時買賣超
        if not both_stocks_df.empty:
            both_stocks_df.to_excel(writer, sheet_name='同時買賣超', index=False)

        # 工作表 4: 每日買超前 50
        sorted_dates = sorted(daily_buy_sell_data.keys(), reverse=True)
        all_daily_buy = []

        for date in sorted_dates:
            if 'buy' in daily_buy_sell_data[date]:
                df = daily_buy_sell_data[date]['buy'].copy()
                df.insert(0, '日期', date)
                all_daily_buy.append(df)

        if all_daily_buy:
            combined_buy = pd.concat(all_daily_buy, ignore_index=True)
            combined_buy.to_excel(writer, sheet_name='每日買超前50', index=False)

        # 工作表 5: 每日賣超前 50
        all_daily_sell = []

        for date in sorted_dates:
            if 'sell' in daily_buy_sell_data[date]:
                df = daily_buy_sell_data[date]['sell'].copy()
                df.insert(0, '日期', date)
                all_daily_sell.append(df)

        if all_daily_sell:
            combined_sell = pd.concat(all_daily_sell, ignore_index=True)
            combined_sell.to_excel(writer, sheet_name='每日賣超前50', index=False)

        # 工作表 6: ETF 買超前 50
        sorted_etf_dates = sorted(etf_daily_data.keys(), reverse=True)
        all_etf_daily = []

        for date in sorted_etf_dates:
            if 'buy' in etf_daily_data[date]:
                df = etf_daily_data[date]['buy'].copy()
                df.insert(0, '日期', date)
                all_etf_daily.append(df)

        if all_etf_daily:
            combined_etf = pd.concat(all_etf_daily, ignore_index=True)
            combined_etf.to_excel(writer, sheet_name='ETF買超前50', index=False)

        # 工作表 7: 新進買超榜
        if new_buy_stocks:
            new_buy_data = []

            latest_buy_df = daily_buy_sell_data.get(latest_date, {}).get('buy')

            if latest_buy_df is not None:
                for stock_code in new_buy_stocks:
                    stock_df = latest_buy_df[latest_buy_df['證券代號'] == stock_code]

                    if not stock_df.empty:
                        new_buy_data.append(stock_df.iloc[0].to_dict())

            if new_buy_data:
                new_buy_df = pd.DataFrame(new_buy_data)
                new_buy_df = new_buy_df.sort_values('三大法人買賣超股數', ascending=False)
                new_buy_df.to_excel(writer, sheet_name='新進買超榜', index=False)

        # 工作表 8: 新進賣超榜
        if new_sell_stocks:
            new_sell_data = []

            latest_sell_df = daily_buy_sell_data.get(latest_date, {}).get('sell')

            if latest_sell_df is not None:
                for stock_code in new_sell_stocks:
                    stock_df = latest_sell_df[latest_sell_df['證券代號'] == stock_code]

                    if not stock_df.empty:
                        new_sell_data.append(stock_df.iloc[0].to_dict())

            if new_sell_data:
                new_sell_df = pd.DataFrame(new_sell_data)
                new_sell_df = new_sell_df.sort_values('三大法人買賣超股數')
                new_sell_df.to_excel(writer, sheet_name='新進賣超榜', index=False)

        # 工作表 9: 觀察買超股
        if observable_buy_stocks:
            observable_buy_data = []

            latest_file_df = daily_buy_sell_data.get(latest_date, {}).get('buy')

            if latest_file_df is not None:
                for stock_code in observable_buy_stocks:
                    stock_df = latest_file_df[latest_file_df['證券代號'] == stock_code]

                    if not stock_df.empty:
                        observable_buy_data.append(stock_df.iloc[0].to_dict())

            if observable_buy_data:
                observable_buy_df = pd.DataFrame(observable_buy_data)
                observable_buy_df = observable_buy_df.sort_values('三大法人買賣超股數', ascending=False)
                observable_buy_df.to_excel(writer, sheet_name='觀察買超股', index=False)

        # 工作表 10: 觀察賣超股
        if observable_sell_stocks:
            observable_sell_data = []

            latest_file_df = daily_buy_sell_data.get(latest_date, {}).get('sell')

            if latest_file_df is not None:
                for stock_code in observable_sell_stocks:
                    stock_df = latest_file_df[latest_file_df['證券代號'] == stock_code]

                    if not stock_df.empty:
                        observable_sell_data.append(stock_df.iloc[0].to_dict())

            if observable_sell_data:
                observable_sell_df = pd.DataFrame(observable_sell_data)
                observable_sell_df = observable_sell_df.sort_values('三大法人買賣超股數')
                observable_sell_df.to_excel(writer, sheet_name='觀察賣超股', index=False)

def beautify_excel(file_path):
    """美化 Excel"""
    wb = load_workbook(file_path)

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)


# ============================================================================
# 第三步:圖表生成程式
# ============================================================================

class Config:
    """配置管理類別"""

    OVERWRITE_EXISTING = True
    MARKET_TYPE = 'TSE'
    RUN_ALL = True

    FONT_PATH = None

    @staticmethod
    def setup_config(market_type='TSE'):
        """設定所有路徑變數"""
        base_path = BASE_DIR

        if market_type == 'TSE':
            config = {
                'market_type': market_type,
                'market_name': '上市',
                'history_folder': os.path.join(base_path, 'StockHistory'),
                'html_output_folder': os.path.join(base_path, 'StockHTML'),
                'png_output_folder': os.path.join(base_path, 'StockPNG'),
                'stocklist_folder': os.path.join(base_path, 'StockList'),
            }
        else:  # OTC
            config = {
                'market_type': market_type,
                'market_name': '上櫃',
                'history_folder': os.path.join(base_path, 'StockOTCHistory'),
                'html_output_folder': os.path.join(base_path, 'StockOTCHTML'),
                'png_output_folder': os.path.join(base_path, 'StockOTCPNG'),
                'stocklist_folder': os.path.join(base_path, 'StockList'),
            }

        os.makedirs(config['html_output_folder'], exist_ok=True)
        os.makedirs(config['png_output_folder'], exist_ok=True)

        print(f"{'='*80}")
        print(f"市場類型: {market_type} ({config['market_name']})")
        print(f"圖表格式: HTML + PNG (雙格式輸出)")
        print(f"歷史數據資料夾: {config['history_folder']}")
        print(f"HTML輸出資料夾: {config['html_output_folder']}")
        print(f"PNG輸出資料夾: {config['png_output_folder']}")
        print(f"{'='*80}\n")

        return config


class Utils:
    """工具函數類別"""

    @staticmethod
    def setup_chinese_font(base_dir):
        """設定中文字體"""
        font_path = os.path.join(base_dir, 'StockList', 'Font.ttf')

        if os.path.exists(font_path):
            Config.FONT_PATH = font_path
            print(f"✓ 找到字體檔案: {font_path}")
        else:
            print(f"⚠ 找不到字體檔案: {font_path}")
            print("  HTML 圖表將使用預設字體")
            Config.FONT_PATH = None

        return Config.FONT_PATH

    @staticmethod
    def read_csv_auto_encoding(file_path):
        """自動偵測編碼讀取 CSV"""
        encodings = ['utf-8-sig', 'utf-8', 'cp950', 'big5']

        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                return df
            except:
                continue

        raise ValueError(f"無法讀取檔案: {file_path}")


class ChartGenerator:
    """圖表生成類別"""

    @staticmethod
    def create_chart(stock_code, stock_name, df, sector=""):
        """建立互動式圖表"""
        if df.empty or len(df) < 2:
            print(f"  ⚠️  資料不足,無法生成圖表")
            return None

        fig = make_subplots(
            rows=2, cols=1,
            row_heights=[0.7, 0.3],
            vertical_spacing=0.08,
            subplot_titles=(
                f'{stock_code} {stock_name} - 三大法人買賣超 & 股價走勢',
                '三大法人買賣超量'
            ),
            specs=[[{"secondary_y": True}], [{"secondary_y": False}]]
        )

        # 圖 1: 股價線圖
        fig.add_trace(
            go.Scatter(
                x=df['日期'],
                y=df['收盤價'],
                name='收盤價',
                line=dict(color='#2E86DE', width=2),
                mode='lines+markers',
                marker=dict(size=4),
                hovertemplate='<b>日期</b>: %{x}<br><b>收盤價</b>: %{y:.2f}<extra></extra>'
            ),
            row=1, col=1, secondary_y=False
        )

        # 圖 1: 買賣超柱狀圖
        colors = ['#10AC84' if x > 0 else '#EE5A6F' for x in df['三大法人買賣超股數']]

        fig.add_trace(
            go.Bar(
                x=df['日期'],
                y=df['三大法人買賣超股數'],
                name='買賣超',
                marker_color=colors,
                opacity=0.6,
                hovertemplate='<b>日期</b>: %{x}<br><b>買賣超</b>: %{y:,}<extra></extra>'
            ),
            row=1, col=1, secondary_y=True
        )

        # 圖 2: 買賣超柱狀圖
        fig.add_trace(
            go.Bar(
                x=df['日期'],
                y=df['三大法人買賣超股數'],
                name='買賣超',
                marker_color=colors,
                showlegend=False,
                hovertemplate='<b>日期</b>: %{x}<br><b>買賣超</b>: %{y:,}<extra></extra>'
            ),
            row=2, col=1
        )

        # 更新 Y 軸
        fig.update_yaxes(title_text="股價 (元)", row=1, col=1, secondary_y=False)
        fig.update_yaxes(title_text="買賣超 (張)", row=1, col=1, secondary_y=True)
        fig.update_yaxes(title_text="買賣超 (張)", row=2, col=1)

        # 更新 X 軸
        fig.update_xaxes(title_text="日期", row=2, col=1)

        # 整體佈局
        fig.update_layout(
            height=900,
            hovermode='x unified',
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            ),
            margin=dict(l=80, r=80, t=100, b=80)
        )

        return fig


class HtmlToPng:
    """HTML 轉 PNG 類別"""

    _driver = None

    @classmethod
    def get_driver(cls):
        """取得或建立 WebDriver"""
        if cls._driver is None:
            try:
                from selenium import webdriver
                from selenium.webdriver.chrome.options import Options

                chrome_options = Options()
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--disable-dev-shm-usage')
                chrome_options.add_argument('--disable-gpu')
                chrome_options.add_argument('--window-size=1920,1080')

                cls._driver = webdriver.Chrome(options=chrome_options)
                print("✓ WebDriver 初始化成功")

            except Exception as e:
                print(f"⚠️  WebDriver 初始化失敗: {e}")
                cls._driver = None

        return cls._driver

    @classmethod
    def cleanup(cls):
        """清理 WebDriver"""
        if cls._driver:
            try:
                cls._driver.quit()
                print("✓ WebDriver 已關閉")
            except:
                pass
            cls._driver = None

    @staticmethod
    def convert(html_path, png_path):
        """將 HTML 轉換為 PNG"""
        driver = HtmlToPng.get_driver()

        if driver is None:
            return False

        try:
            driver.get(f'file://{html_path}')
            time.sleep(2)

            driver.save_screenshot(png_path)
            return True

        except Exception as e:
            print(f"  ⚠️  轉換失敗: {e}")
            return False


class Processor:
    """處理器類別"""

    @staticmethod
    def process_stock(stock_code, base_dir, config):
        """處理單一股票"""
        print(f"\n處理股票: {stock_code}")

        csv_path = os.path.join(config['history_folder'], f"{stock_code}.csv")

        if not os.path.exists(csv_path):
            print(f"  ⚠️  找不到歷史資料: {csv_path}")
            return None

        html_output = os.path.join(config['html_output_folder'], f"{stock_code}.html")
        png_output = os.path.join(config['png_output_folder'], f"{stock_code}.png")

        if not Config.OVERWRITE_EXISTING:
            if os.path.exists(html_output) and os.path.exists(png_output):
                print(f"  ⊙ 檔案已存在,跳過")
                return None

        try:
            df = Utils.read_csv_auto_encoding(csv_path)

            if df.empty:
                print(f"  ⚠️  CSV 檔案為空")
                return False

            if '日期' not in df.columns:
                print(f"  ⚠️  缺少'日期'欄位")
                return False

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
            df = df.dropna(subset=['日期'])

            if df.empty:
                print(f"  ⚠️  日期轉換後無有效資料")
                return False

            df = df.sort_values('日期')

            if '三大法人買賣超股數' in df.columns:
                df['三大法人買賣超股數'] = pd.to_numeric(
                    df['三大法人買賣超股數'].astype(str).str.replace(',', ''),
                    errors='coerce'
                )

            if '收盤價' in df.columns:
                df['收盤價'] = pd.to_numeric(df['收盤價'], errors='coerce')

            stock_name = ""
            if '證券名稱' in df.columns and not df.empty:
                stock_name = df['證券名稱'].iloc[0]

            sector = ""

            fig = ChartGenerator.create_chart(stock_code, stock_name, df, sector)

            if fig is None:
                return False

            fig.write_html(html_output)
            print(f"  ✓ HTML 已儲存: {os.path.basename(html_output)}")

            success = HtmlToPng.convert(html_output, png_output)

            if success:
                print(f"  ✓ PNG 已儲存: {os.path.basename(png_output)}")
                return True
            else:
                print(f"  ⚠️  PNG 轉換失敗")
                return True

        except Exception as e:
            print(f"  ❌ 處理失敗: {e}")
            return False

    @staticmethod
    def batch_process_all_stocks(base_dir, config):
        """批次處理所有股票"""
        print(f"\n{'='*80}")
        print(f"批次處理模式 - {config['market_name']}")
        print(f"{'='*80}\n")

        history_folder = config['history_folder']

        if not os.path.exists(history_folder):
            print(f"❌ 歷史資料夾不存在: {history_folder}")
            return

        csv_files = glob.glob(os.path.join(history_folder, '*.csv'))

        if not csv_files:
            print(f"❌ 找不到歷史資料檔案")
            return

        stock_codes = [os.path.basename(f).replace('.csv', '') for f in csv_files]
        stock_codes = sorted(stock_codes)

        if not stock_codes:
            print("❌ 無法取得股票清單")
            return

        print(f"✓ 找到 {len(stock_codes)} 支股票")

        success_count = 0
        fail_count = 0
        skip_count = 0

        start_time = datetime.now()

        for idx, stock_code in enumerate(stock_codes, 1):
            print(f"\n{'='*70}")
            print(f"進度: [{idx}/{len(stock_codes)}] ({idx/len(stock_codes)*100:.1f}%)")
            print(f"{'='*70}")

            result = Processor.process_stock(stock_code, base_dir, config)

            if result is True:
                success_count += 1
            elif result is False:
                fail_count += 1
            elif result is None:
                skip_count += 1

        end_time = datetime.now()
        elapsed_time = (end_time - start_time).total_seconds()

        print("\n" + "="*70)
        print("批次處理完成")
        print("="*70)
        print(f"總股票數: {len(stock_codes)}")
        print(f"成功處理: {success_count}")
        print(f"跳過處理: {skip_count}")
        print(f"處理失敗: {fail_count}")
        print(f"處理時間: {elapsed_time:.1f} 秒 ({elapsed_time/60:.1f} 分鐘)")
        print("="*70)

        HtmlToPng.cleanup()


# ============================================================================
# 主程式執行函數
# ============================================================================

def run_step1_crawler():
    """執行第一步:爬蟲程式"""
    print("\n" + "🔥"*40)
    print("第一步:執行爬蟲程式")
    print("🔥"*40 + "\n")
    
    start_date = datetime(2025, 1, 1)
    end_date = datetime.now()
    print(f"日期範圍: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
    print(f"儲存位置: {BASE_DIR}/")
    print()

    start_time = time.time()
    dirs = {
        'StockDaily': os.path.join(BASE_DIR, 'StockDaily'),
        'StockShares': os.path.join(BASE_DIR, 'StockShares'),
        'StockOTCDaily': os.path.join(BASE_DIR, 'StockOTCDaily'),
        'StockOTCShares': os.path.join(BASE_DIR, 'StockOTCShares')
    }

    results = {}
    results['twse_daily'] = crawl_twse_daily(start_date, end_date, dirs['StockDaily'])
    results['twse_inst'] = crawl_twse_institutional(start_date, end_date, dirs['StockShares'])
    results['otc_daily'] = crawl_otc_daily(start_date, end_date, dirs['StockOTCDaily'])
    results['otc_inst'] = crawl_otc_institutional(start_date, end_date, dirs['StockOTCShares'])

    elapsed_time = time.time() - start_time
    print("="*60)
    print("📊 第一步執行結果摘要")
    print("="*60)
    print(f"✓ 上市每日交易:  {results['twse_daily']} 個檔案")
    print(f"✓ 上市三大法人:  {results['twse_inst']} 個檔案")
    print(f"✓ 上櫃每日交易:  {results['otc_daily']} 個檔案")
    print(f"✓ 上櫃三大法人:  {results['otc_inst']} 個檔案")
    print("-"*60)
    print(f"總計下載:{sum(results.values())} 個檔案")
    print(f"執行時間:{elapsed_time:.1f} 秒")
    print("="*60)

def run_step2_analysis(market_type='TSE'):
    """執行第二步:分析程式"""
    print(f"\n{'🔥'*40}")
    print(f"第二步分析:{market_type} ({'上市' if market_type == 'TSE' else '上櫃'})")
    print(f"{'🔥'*40}\n")
    
    config = setup_config(market_type=market_type)
    allowed_stock_codes, stock_sector_map, etf_stock_codes = load_stock_list(config['market_list_path'])
    stock_daily_prices = load_stock_daily_prices(config['stock_daily_folder'], allowed_stock_codes)
    latest_61_files = get_latest_files(config['folder_path'], num_files=61)
    
    (all_data, daily_buy_sell_data, etf_daily_data, buy_top20_tracker,
     sell_top20_tracker, daily_buy_stocks, daily_sell_stocks,
     daily_all_stocks, all_historical_data, statistics) = process_shares_files(
        latest_61_files, allowed_stock_codes, stock_daily_prices,
        stock_sector_map, etf_stock_codes
    )
    
    stock_statistics = calculate_stock_statistics(all_historical_data, config['sigma_threshold'])
    
    (new_buy_stocks, new_sell_stocks, observable_buy_stocks, observable_sell_stocks,
     latest_date, latest_buy_stocks_50, latest_sell_stocks_50) = analyze_new_entries_and_observables(
        latest_61_files[0], daily_buy_stocks, daily_sell_stocks,
        daily_all_stocks, stock_statistics, allowed_stock_codes,
        config['sigma_threshold']
    )
    
    collect_stock_history(latest_buy_stocks_50, config['folder_path'],
                          config['stock_daily_folder'], config['history_folder'],
                          allowed_stock_codes)
    
    buy_stocks, sell_stocks, both_stocks_set, both_stocks_df = aggregate_analysis(
        buy_top20_tracker, sell_top20_tracker, stock_sector_map,
        aggregate_threshold=config.get('aggregate_threshold', 10000),
        show_top_n=config.get('show_top_n', None)
    )
    
    if buy_stocks is not None and sell_stocks is not None:
        export_to_excel(config['output_path'], buy_stocks, sell_stocks, both_stocks_set,
                       both_stocks_df, daily_buy_sell_data, etf_daily_data, latest_date,
                       new_buy_stocks, new_sell_stocks, observable_buy_stocks,
                       observable_sell_stocks, stock_sector_map, etf_stock_codes)
        beautify_excel(config['output_path'])
        print(f"\n✓ {market_type} 分析完成")
        print(f"✓ Excel 已儲存: {config['output_path']}")

def run_step3_charts(market_type='TSE'):
    """執行第三步:圖表生成"""
    print(f"\n{'🔥'*40}")
    print(f"第三步圖表:{market_type} ({'上市' if market_type == 'TSE' else '上櫃'})")
    print(f"{'🔥'*40}\n")
    
    config = Config.setup_config(market_type=market_type)
    Utils.setup_chinese_font(BASE_DIR)
    Processor.batch_process_all_stocks(BASE_DIR, config)
    print(f"\n✓ {market_type} 圖表完成")

def main():
    """主程式"""
    print("\n" + "="*80)
    print("台灣股市資料完整處理流程 - GitHub Actions 版")
    print("="*80)
    print("執行環境:", BASE_DIR)
    print("="*80 + "\n")
    
    # 步驟 1:爬蟲
    run_step1_crawler()
    
    # 步驟 2:清理 History
    print("\n" + "🔥"*40)
    print("步驟 2:清理 History 資料夾")
    print("🔥"*40)
    delete_folders(['StockHistory', 'StockOTCHistory'])
    
    # 步驟 3-4:分析
    run_step2_analysis('TSE')
    run_step2_analysis('OTC')
    
    # 步驟 5:清理圖表
    print("\n" + "🔥"*40)
    print("步驟 5:清理圖表資料夾")
    print("🔥"*40)
    delete_folders(['StockHTML', 'StockPNG', 'StockOTCHTML', 'StockOTCPNG'])
    
    # 步驟 6-7:圖表
    run_step3_charts('TSE')
    run_step3_charts('OTC')
    
    # 完成
    print("\n" + "🎉"*40)
    print("所有流程已完成!")
    print("🎉"*40 + "\n")

if __name__ == "__main__":
    main()
